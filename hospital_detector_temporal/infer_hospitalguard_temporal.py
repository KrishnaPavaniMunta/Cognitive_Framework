"""
infer_hospitalguard_temporal.py  —  Short-Term Object Tracking
─────────────────────────────────────────────────────────────────────────────
HospitalGuard-109  —  interactive inference with Grounding DINO fallback
                       and ByteTrack short-term object tracking.

Architecture
────────────
  Layer 1  YOLO V1+V3 ensemble (every frame, all 109 classes)
             V1 (106-class): all COCO + hospital classes
             V3 (109-class): bag / exit_sign / spillage + overlap NMS
  Layer 2  Grounding DINO fallback (fires every DINO_VIDEO_INTERVAL_SEC
           seconds, only for weak classes YOLO found zero boxes for)
  Layer 3  ByteTrack short-term tracker (every frame)
             Assigns persistent track IDs so each object keeps the same
             ID label across frames for as long as it remains visible.

Tracking duration
─────────────────
  Occlusion tolerance  : 3 seconds  (lost_track_buffer = round(fps * 3))
    — if an object disappears for ≤ 3 s it re-emerges with the SAME ID.
    — beyond 3 s the track is dropped and a new ID is assigned on re-entry.
  Room-state memory    : ~5 seconds  (_ROOM_STATE_STALE_FRAMES = 150 @ 30 fps)
    — the last known position / confidence of every track is kept in
      room_state for 5 s after it was last detected, enabling flicker-free
      PPE alerts even during brief occlusions beyond the tracker window.

Weak-class DINO targets  (ensemble AP50 < 0.25)
────────────────────────
  surgical_scissor     glove         mask           hair_net
  surgical_light       radiator      iv_stand       exit_sign
  medical_tray         infusion_pump

Output
──────
  Annotated video  →  outputs/hospitalguard_output/<stem>_<ts>.mp4
  Annotated image  →  outputs/hospitalguard_output/<stem>_<ts>.jpg
  Excel log        →  outputs/hospitalguard_log.xlsx
    YOLO+tracked boxes : supervision default colour palette
    DINO+tracked boxes : orange with [DINO] prefix on label
    All boxes labelled : "<class> #<track_id>  <conf%>"

Usage
─────
  python infer_hospitalguard_temporal.py
  (paste image or video URL when prompted; type quit to exit)
"""

import os
os.environ.setdefault("PYTORCH_ALLOC_CONF", "expandable_segments:True")
import tempfile
import requests
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import cv2
import torch
import numpy as np
import supervision as sv
from PIL import Image
from ultralytics import YOLO
from transformers import AutoProcessor, AutoModelForZeroShotObjectDetection

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.resolve()
ROOT_DIR   = BASE_DIR.parent   # yolo_tr/ — where outputs/ and models/ live
V1_PATH    = ROOT_DIR / "outputs/runs/hospital/phase2_neck_head/weights/best.pt"
V3_PATH    = ROOT_DIR / "outputs/runs/hospital_v3/phase2_neck_head/weights/best.pt"
OUT_DIR    = ROOT_DIR / "outputs/hospitalguard_output"
EXCEL_PATH = ROOT_DIR / "outputs/hospitalguard_log.xlsx"
OUTPUT_RUN_TAG = "upd_motion_stab_v1"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── YOLO settings ──────────────────────────────────────────────────────────────
CONF            = 0.25
IOU             = 0.45
LOW_CONF_THRESH = 0.50

# ── V3 ensemble routing ────────────────────────────────────────────────────────
# Both V1 and V3 actively detect these — pool then NMS
V3_WORKING_OVERLAP = {"wheelchair", "door", "fire_extinguisher"}
# Only V3 vocabulary contains these — use V3 boxes alone
V3_ONLY_NEW        = {"bag", "exit_sign", "spillage"}

# ── Media type helpers ────────────────────────────────────────────────────────
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff"}
VIDEO_EXTS = {".mp4", ".avi", ".mov", ".mkv", ".webm", ".m4v"}

# ── Grounding DINO config ──────────────────────────────────────────────────────
DINO_MODEL_ID        = "IDEA-Research/grounding-dino-base"
DINO_DEVICE          = "cuda"
DINO_TEXT_THR        = 0.25   # text-alignment threshold (shared)
DINO_VIDEO_INTERVAL_SEC = 1.0   # motion stabilizer horizon (~15 frames at 15 FPS)
DINO_VIDEO_INTERVAL_FRAMES = 15  # run DINO once every 15 frames

# Hospital-relevant weak classes only (AP50 < 0.25 or user-requested).
# COCO-only classes that never appear in hospitals (hair drier, toothbrush,
# hot dog) are excluded — they only generate false positives.
# Format: canonical YOLO name → (elaborated DINO prompt, per-class box threshold)
DINO_FALLBACK: dict[str, tuple[str, float]] = {
    # ── OR / surgical scene ───────────────────────────────────────────────────
    # surgical_scissor: new prompt improves DINO-base score 0.275→0.505 vs old prompt.
    "surgical_scissor": ("surgical scissors. stainless steel. instrument tray.", 0.30),
    # surgical_light (AP50=0.732) scores near-zero on OR images despite strong validation
    # numbers — domain shift in lighting/angle. DINO-base reliably detects it (0.508).
    "surgical_light":   ("large round overhead surgical operating light.", 0.40),
    # ── PPE — mask, glove, hair_net ──────────────────────────────────────────
    # glove (AP50=0.830) also scores 0 YOLO in OR scene; DINO-base gets 0.303.
    # Context-gated to clinical scenes to prevent baseball-glove FPs.
    "glove":            ("blue surgical latex glove. rubber medical glove.", 0.28),
    # mask (AP50=0.689): worn on face; DINO isolates it from non-medical face coverings
    # via the prompt. Context-gated to scenes with healthcare_worker/patient.
    "mask":             ("surgical face mask. blue medical face mask worn by clinician.", 0.35),
    # hair_net (AP50=0.637): worn on head in OR/wards. Prone to FP with any headwear
    # so context-gated strictly to clinical anchors.
    "hair_net":         ("surgical hair net. disposable hair cover worn in operating room.", 0.38),
    # ── Infrastructure / hazards ──────────────────────────────────────────────
    # radiator: YOLO AP50=0.981 in val but zero activation on real ward images
    # (315 annotations / 34K training images = 0.93% frequency; severe domain shift).
    # Restored to DINO fallback. Threshold raised to 0.50 to reduce home-scene FPs.
    "radiator":         ("radiator heater mounted on wall.", 0.50),
    # iv_stand (AP50=0.180): worst-performing class in the whole model. Thin vertical
    # pole is confused with mic stands, coat racks, lamps. DINO-base gets 0.45+ on
    # real IV poles when context-gated to clinical scenes.
    "iv_stand":         ("intravenous IV drip stand. metal pole with wheels and hanging bag hook.", 0.40),
    # exit_sign (AP50=0.797 V3 val): safety-critical anchor for egress monitoring.
    # DINO fallback catches signs in dim/angled corridors that confuse YOLO.
    "exit_sign":        ("green rectangular exit sign. illuminated emergency exit sign on wall.", 0.42),
    # iv_stand removed from YOLO-SAHI comment kept for reference:
    # DINO-tiny cannot distinguish IV poles from mic stands, coat racks,
    # floor lamps (all score 0.35–0.54 with any pole-based prompt). YOLO-only covers it.
    # handbag removed: YOLO native class covers it well enough; DINO FPs in non-hospital scenes
    # knife removed: DINO-tiny cannot distinguish scalpel from kitchen knife
    # at the prompt level — YOLO's native knife class (AP50=0.200) covers it.
    # test_tube removed from here — handled by DINO_SAHI (SAHI + neg prompting).
    # ── DINO-only classes (not in any YOLO model) ────────────────────────────
    # medical_tray: no YOLO class exists for this. DINO fires whenever OR/procedure
    # context is confirmed. Tray content is irrelevant — detects the tray itself.
    "medical_tray":     ("stainless steel medical tray. rectangular or oval silver tray used in clinical or hospital setting.", 0.40),
    # infusion_pump: YOLO class exists (AP50 reasonable in val) but misses in
    # real ward images due to device variety / domain shift. DINO fallback.
    "infusion_pump":    ("infusion pump. intravenous IV pump with digital display mounted on pole.", 0.38),
}

# Classes that must be queried in isolation — they compete with other prompts
# on shared visual features (e.g. surgical_scissor vs knife both have blades,
# glove vs baseball-glove share the word 'glove',
# mask vs non-medical face covering, hair_net vs regular hat).
DINO_ISOLATED: set[str] = {"surgical_scissor", "glove", "mask", "hair_net", "iv_stand", "medical_tray"}

# Classes that use SAHI (sliced inference) + negative prompting.
# SAHI slices large images into 640-px patches so small objects fill more of
# the frame; negative prompting forces DINO to label confusable objects
# (syringes, vials, pens) before the target phrase, suppressing FPs.
# Format: canonical name → {phrase, distractors, threshold, slice_size, overlap}
DINO_SAHI: dict[str, dict] = {
    "test_tube": {
        "phrase":       "glass test tube. coloured rubber cap.",
        "distractors":  ["syringe", "vial", "glass bottle", "glass jar",
                         "pen", "marker", "kitchen knife", "food", "cutlery"],
        "threshold":    0.44,
        "slice_size":   640,
        "overlap":      0.25,
    },
}

# Classes where YOLO SAHI supplement runs at image-inference time.
# Small objects that fill too few pixels at full resolution become clearly
# visible when YOLO runs on a 640-px crop of just that region.
# Only used in run_image() — video keeps every-frame full-resolution YOLO
# to maintain throughput.
YOLO_SAHI_CLASSES: set[str] = {"fire_extinguisher", "surgical_scissor", "iv_stand"}
YOLO_SAHI_SLICE   = 640   # patch edge length in pixels
YOLO_SAHI_OVERLAP = 0.25  # fraction overlap between adjacent patches

# Context anchors for DINO gating (two-step contextual verification).
# DINO will only fire for a class if at least one of its anchor classes was
# already detected by YOLO. This prevents DINO from hallucinating in scenes
# that structurally cannot contain the target object.
DINO_CONTEXT_ANCHORS: dict[str, set[str]] = {
    # surgical_scissor: misplaced scissors are hazardous. Fire when confirmed
    # clinical/OR scene via PPE or equipment (not surgical_light — too specific;
    # not hair_net/patient — too broad for OR-only inference).
    "surgical_scissor": {
        "healthcare_worker", "glove", "mask",
        "hospital_bed", "infusion_pump", "iv_bag",
    },
    # surgical_light: wall-mounted overhead lamp. Gate on bed/IV equipment
    # which co-occur with OR lights. Removed glove/mask — too generic.
    "surgical_light": {
        "healthcare_worker", "hair_net", "monitor_hosp",
        "hospital_bed", "patient", "infusion_pump", "iv_bag",
    },
    "glove": {
        "surgical_light", "healthcare_worker", "hair_net", "mask",
        "hospital_bed", "patient", "infusion_pump", "monitor_hosp",
    },
    "mask": {
        "healthcare_worker", "hair_net", "glove", "surgical_light",
        "hospital_bed", "patient", "infusion_pump",
    },
    # hair_net: fire ONLY when a healthcare_worker is present or a clinical
    # patient/bed context is confirmed. Prevents FP on hats/bonnets.
    "hair_net": {
        "healthcare_worker", "hospital_bed", "patient",
    },
    # radiator removed: YOLO AP50=0.981, relying on YOLO only for now.
    # iv_stand: strict clinical gate — pole-like objects are ubiquitous.
    # Only fire when IV equipment (bag/pump) or patient context confirmed.
    "iv_stand": {
        "iv_bag", "infusion_pump", "patient", "hospital_bed",
    },
    # exit_sign: safety-critical egress marker. Only fire when a door is
    # detected — signs without doors are likely decorative/non-egress.
    "exit_sign": {
        "door",
    },
    # test_tube: biohazard. Gate on IV/monitor equipment only — removed
    # healthcare_worker/hair_net/patient which were too loose.
    "test_tube": {
        "iv_bag", "infusion_pump", "monitor_hosp",
        "hospital_bed", "patient_monitor", "glove", "mask",
        "cabinet", "nasal_cannula", "surgical_light",
    },
    # medical_tray: DINO-only class. Fire when OR/procedure scene confirmed.
    "medical_tray": {
        "surgical_scissor", "glove", "mask",
        "healthcare_worker", "surgical_light", "hospital_bed",
    },
    # infusion_pump: fire only when patient/IV context confirmed.
    "infusion_pump": {
        "iv_bag", "iv_stand", "patient", "hospital_bed", "healthcare_worker",
    },
}

# Maximum bounding-box area (as fraction of image area) for DINO detections.
# DINO-base tends to predict large scene-level boxes for some classes;
# this cap discards any box whose area exceeds the given fraction.
# Applies in both _dino_query() and _sahi_dino_query().
DINO_MAX_BOX_FRAC: dict[str, float] = {
    "surgical_scissor": 0.12,   # small instrument
    "surgical_light":   0.30,   # large overhead lamp — still < 30% of frame
    "glove":            0.10,   # hand-sized object
    "mask":             0.08,   # face-sized object
    "hair_net":         0.08,   # head-sized object
    "iv_stand":         0.25,   # tall pole — narrow but tall
    "exit_sign":        0.10,   # wall-mounted sign
    "test_tube":        0.06,   # tiny object
    "radiator":         0.20,   # wall panel — reasonably large but not full-wall
    "medical_tray":     0.20,   # tray can fill a significant portion of frame if close
    "infusion_pump":    0.15,   # pump box ~ bedside device size
}

# YOLO bottle suppression in confirmed surgical scenes.
# Requires a quorum of OR-specific indicators to avoid accidental suppression.
BOTTLE_SURGICAL_SUPPRESSOR = {"surgical_light", "glove", "mask", "hair_net", "healthcare_worker"}
BOTTLE_SURGICAL_QUORUM     = 2

# ── Stable class-name → int mapping (YOLO 109 + DINO-only classes) ─────────────
# Used by ByteTrack bridge to build sv.Detections with consistent class_id values.
# Sorted alphabetically so IDs never shift when classes are added to DINO_FALLBACK.
CLASS_NAMES: list[str] = sorted([
    'airplane', 'apple', 'backpack', 'bag', 'banana', 'baseball bat',
    'baseball glove', 'bathroom_labels', 'bear', 'bed', 'bench', 'bench_hosp',
    'bicycle', 'bird', 'boat', 'book', 'bottle', 'bowl', 'broccoli', 'bus',
    'cabinet', 'cake', 'car', 'carrot', 'cat', 'cell phone', 'chair', 'clock',
    'couch', 'cow', 'cup', 'dining table', 'dog', 'donut', 'door', 'elephant',
    'exit_sign', 'fire hydrant', 'fire_extinguisher', 'fork', 'frisbee',
    'giraffe', 'glove', 'hair drier', 'hair_net', 'handbag', 'healthcare_worker',
    'horse', 'hospital_bed', 'hospital_stretcher', 'hot dog', 'infusion_pump',
    'iv_bag', 'iv_stand', 'keyboard', 'kite', 'knife', 'laptop', 'mask',
    'medical_tray',   # DINO-only — not in any YOLO model
    'microwave', 'monitor_hosp', 'motorcycle', 'mouse', 'nasal_cannula',
    'orange', 'oven', 'parking meter', 'patient', 'patient_monitor', 'person',
    'pizza', 'potted plant', 'radiator', 'reception_counter', 'refrigerator',
    'remote', 'sandwich', 'scissors', 'security_camera', 'sheep', 'sink',
    'skateboard', 'skis', 'snowboard', 'spillage', 'spoon', 'sports ball',
    'stop sign', 'suitcase', 'surfboard', 'surgical_light', 'surgical_scissor',
    'teddy bear', 'tennis racket', 'test_tube', 'tie', 'toaster', 'toilet',
    'toothbrush', 'traffic light', 'train', 'truck', 'tv', 'umbrella', 'vase',
    'vending_machines', 'wheelchair', 'wine glass', 'zebra',
])
CLASS_TO_ID: dict[str, int] = {name: i for i, name in enumerate(CLASS_NAMES)}

# Normalize ambiguous/general COCO labels to hospital-specific classes.
CLASS_ALIASES: dict[str, str] = {
    "tv": "monitor_hosp",
    "bed": "hospital_bed",
}


def _canonical_class_name(name: str) -> str:
    return CLASS_ALIASES.get(name, name)

# ── Excel layout (mirrors infer_v3.py / infer_ensemble.py) ────────────────────
FIXED_HEADERS = [
    "Image ID",
    "Ground Truth (Target)",
    "Detected?",
    "Confidence Score",
    "Result Type",
    "Notes (False Positives/Misses)",
]
NUM_FIXED    = len(FIXED_HEADERS)
FIXED_WIDTHS = {"A": 55, "B": 24, "C": 12, "D": 22, "E": 22, "F": 60}

RESULT_COLORS = {
    "TP":                  "C6EFCE",
    "TP (Low Conf)":       "FFEB9C",
    "FN (False Negative)": "FFC7CE",
    "FP (False Positive)": "FFCC99",
    "TN (True Negative)":  "D9E1F2",
}
HEADER_FILL_FIXED = PatternFill("solid", fgColor="1D1E37")
HEADER_FILL_CLASS = PatternFill("solid", fgColor="FA643F")
HEADER_FONT       = Font(bold=True, color="FFFFFF")


# ══════════════════════════════════════════════════════════════════════════════
# Layer 1 — YOLO V1+V3 ensemble
# ══════════════════════════════════════════════════════════════════════════════

def _nms_merge(combined: list, iou_thresh: float = 0.45) -> list:
    """NMS over a list of (x1, y1, x2, y2, conf) tuples."""
    if not combined:
        return []
    from torchvision.ops import nms as tv_nms
    boxes  = torch.tensor([[d[0], d[1], d[2], d[3]] for d in combined], dtype=torch.float32)
    scores = torch.tensor([d[4] for d in combined], dtype=torch.float32)
    kept   = tv_nms(boxes, scores, iou_thresh).tolist()
    return [combined[i] for i in kept]


def ensemble_infer(v1: YOLO, v3: YOLO, img_path: Path) -> dict:
    """
    Run V1 + V3 YOLO ensemble.
    Returns {class_name: [(x1, y1, x2, y2, conf), ...]}
    """
    r1 = v1(str(img_path), conf=CONF, iou=IOU, verbose=False)[0]
    r3 = v3(str(img_path), conf=CONF, iou=IOU, verbose=False)[0]

    v1_dets: dict[str, list] = defaultdict(list)
    if r1.boxes is not None:
        for box in r1.boxes:
            name = _canonical_class_name(v1.names[int(box.cls)])
            xyxy = box.xyxy[0].cpu().tolist()
            v1_dets[name].append((*xyxy, float(box.conf)))

    v3_dets: dict[str, list] = defaultdict(list)
    if r3.boxes is not None:
        for box in r3.boxes:
            name = _canonical_class_name(v3.names[int(box.cls)])
            xyxy = box.xyxy[0].cpu().tolist()
            v3_dets[name].append((*xyxy, float(box.conf)))

    merged: dict[str, list] = {}

    # V3-exclusive classes — V3 is the only model that knows them
    for cls in V3_ONLY_NEW:
        if cls in v3_dets:
            merged[cls] = v3_dets[cls]

    # Overlap classes — combine both models then de-duplicate via NMS
    for cls in V3_WORKING_OVERLAP:
        combined = v1_dets.get(cls, []) + v3_dets.get(cls, [])
        if combined:
            merged[cls] = _nms_merge(combined, IOU)

    # All remaining V1 classes
    exclude = V3_WORKING_OVERLAP | V3_ONLY_NEW
    for cls, dets in v1_dets.items():
        if cls not in exclude:
            merged[cls] = dets

    return _post_filter_yolo(merged)


def _post_filter_yolo(merged: dict) -> dict:
    """
    Class-specific post-processing filters applied to every YOLO output.
      bottle — remove boxes wider than 150 px: instrument trays and shelves
               are misclassified as bottles; real bottles are narrow.
      bag    — require confidence ≥ 0.35: surgical drapes score ~0.28 while
               real bags (patient bags, IV bags) score 0.85+.
    """
    if "bottle" in merged:
        merged["bottle"] = [d for d in merged["bottle"] if (d[2] - d[0]) <= 150]
        if not merged["bottle"]:
            del merged["bottle"]
    if "bag" in merged:
        merged["bag"] = [d for d in merged["bag"] if d[4] >= 0.35]
        if not merged["bag"]:
            del merged["bag"]
    return merged


def _yolo_sahi_supplement(v1: YOLO, v3: YOLO, img_path: Path) -> dict:
    """
    Run YOLO V1+V3 on overlapping 640-px patches (SAHI) for YOLO_SAHI_CLASSES.

    A fire extinguisher on a distant wall, surgical scissors on a tray, or an
    IV stand pole all shrink to very few pixels at full-image resolution, which
    pushes YOLO confidence below the threshold. Inside a 640-px crop of just
    that region the object is much larger and YOLO detects it confidently.

    Strategy:
      1. Slice the image into overlapping YOLO_SAHI_SLICE-px patches.
      2. Run V1 + V3 on each patch in BGR numpy format.
      3. Keep only YOLO_SAHI_CLASSES detections.
      4. Remap patch-local boxes to original image coordinates.
      5. NMS-merge across all patches (IoU > IOU threshold).

    Returns {class: [(x1, y1, x2, y2, conf), ...]} for YOLO_SAHI_CLASSES only.
    """
    img  = Image.open(img_path).convert("RGB")
    W, H = img.size
    step = int(YOLO_SAHI_SLICE * (1 - YOLO_SAHI_OVERLAP))
    xs   = sorted(set(
        list(range(0, max(1, W - YOLO_SAHI_SLICE + 1), step))
        + [max(0, W - YOLO_SAHI_SLICE)]
    ))
    ys   = sorted(set(
        list(range(0, max(1, H - YOLO_SAHI_SLICE + 1), step))
        + [max(0, H - YOLO_SAHI_SLICE)]
    ))

    raw: dict[str, list] = defaultdict(list)
    for x0 in xs:
        for y0 in ys:
            crop = img.crop((x0, y0, min(x0 + YOLO_SAHI_SLICE, W), min(y0 + YOLO_SAHI_SLICE, H)))
            # Convert RGB→BGR so YOLO sees the same channel order as cv2 frames
            patch = cv2.cvtColor(np.array(crop), cv2.COLOR_RGB2BGR)
            r1 = v1(patch, conf=CONF, iou=IOU, verbose=False)[0]
            r3 = v3(patch, conf=CONF, iou=IOU, verbose=False)[0]
            for model, result in ((v1, r1), (v3, r3)):
                if result.boxes is None:
                    continue
                for box in result.boxes:
                    name = _canonical_class_name(model.names[int(box.cls)])
                    if name not in YOLO_SAHI_CLASSES:
                        continue
                    bx1, by1, bx2, by2 = box.xyxy[0].cpu().tolist()
                    raw[name].append(
                        (bx1 + x0, by1 + y0, bx2 + x0, by2 + y0, float(box.conf))
                    )

    result: dict[str, list] = {}
    for cls, dets in raw.items():
        nms_dets = _nms_merge(dets, IOU)
        if nms_dets:
            result[cls] = nms_dets
    return result


# ══════════════════════════════════════════════════════════════════════════════
# Layer 2 — Grounding DINO fallback (lazy-loaded)
# ══════════════════════════════════════════════════════════════════════════════

_dino_processor = None
_dino_model     = None


def _load_dino() -> None:
    global _dino_processor, _dino_model
    if _dino_model is None:
        print(f"  [DINO] Loading {DINO_MODEL_ID} on {DINO_DEVICE.upper()} (first use)...")
        _dino_processor = AutoProcessor.from_pretrained(DINO_MODEL_ID)
        _dino_model = (
            AutoModelForZeroShotObjectDetection
            .from_pretrained(DINO_MODEL_ID)
            .to(DINO_DEVICE)
        )
        _dino_model.eval()
        print("  [DINO] Ready.")


def _dino_query(pil_image: Image.Image,
                phrase_to_canonical: dict,
                phrase_to_threshold: dict) -> dict:
    """
    Single DINO forward pass for one prompt batch.
    Returns {canonical_name: [(x1, y1, x2, y2, conf), ...]}
    """
    prompt = " . ".join(phrase_to_canonical.keys()) + " ."
    inputs = _dino_processor(
        images=pil_image,
        text=prompt,
        return_tensors="pt",
    ).to(DINO_DEVICE)

    torch.cuda.empty_cache()
    with torch.no_grad():
        outputs = _dino_model(**inputs)

    min_thr = min(phrase_to_threshold.values())
    results = _dino_processor.post_process_grounded_object_detection(
        outputs,
        inputs["input_ids"],
        threshold=min_thr,
        text_threshold=DINO_TEXT_THR,
        target_sizes=[pil_image.size[::-1]],
    )[0]

    boxes  = results["boxes"].cpu().numpy()
    scores = results["scores"].cpu().numpy()
    labels = results["text_labels"]

    dino_dets: dict[str, list] = defaultdict(list)
    # When only one class is in the batch (isolated run) any returned box
    # belongs to that class — DINO's text_labels may drop low-scoring tokens
    # (e.g. "stainless steel" in "stainless steel medical tray"), breaking
    # exact/substring matching.  Falling back to the sole class is safe.
    sole_canonical = next(iter(phrase_to_canonical.values())) if len(phrase_to_canonical) == 1 else None
    sole_phrase    = next(iter(phrase_to_canonical.keys()))   if len(phrase_to_canonical) == 1 else None

    for box, score, label in zip(boxes, scores, labels):
        label_clean = label.strip().lower()
        canonical      = phrase_to_canonical.get(label_clean)
        matched_phrase = label_clean
        if canonical is None:
            # Normalised substring check (strip punctuation for comparison)
            import re as _re
            label_norm = _re.sub(r'[^\w\s]', ' ', label_clean).lower()
            for phrase, name in phrase_to_canonical.items():
                phrase_norm = _re.sub(r'[^\w\s]', ' ', phrase).lower()
                # any word overlap heuristic: check if label words appear in phrase
                label_words = set(label_norm.split())
                phrase_words = set(phrase_norm.split())
                if label_words & phrase_words:  # non-empty intersection
                    canonical      = name
                    matched_phrase = phrase
                    break
        if canonical is None and sole_canonical is not None:
            # Isolated-run fallback: assign to the only class in this batch
            canonical      = sole_canonical
            matched_phrase = sole_phrase
        if canonical is None:
            continue
        cls_thr = phrase_to_threshold.get(matched_phrase, 0.40)
        if float(score) < cls_thr:
            continue
        x1, y1, x2, y2 = box.tolist()
        max_frac = DINO_MAX_BOX_FRAC.get(canonical)
        if max_frac is not None:
            W_img, H_img = pil_image.size
            if (x2 - x1) * (y2 - y1) / (W_img * H_img) > max_frac:
                continue
        dino_dets[canonical].append((x1, y1, x2, y2, float(score)))

    return dict(dino_dets)


def _sahi_dino_query(pil_image: Image.Image, cls: str) -> dict:
    """
    SAHI (sliced) + negative-prompting DINO query for a single class.

    Slices the image into overlapping patches so small objects (e.g. test
    tubes) occupy enough pixels for DINO to see fine-grained features.
    Distractors are included in the prompt but not in the canonical map,
    forcing DINO to assign confusable shapes to distractor labels instead
    of the target — those boxes are then silently discarded.

    Returns {cls: [(x1, y1, x2, y2, conf), ...]}
    """
    cfg         = DINO_SAHI[cls]
    phrase      = cfg["phrase"]
    distractors = cfg["distractors"]
    threshold   = cfg["threshold"]
    slice_size  = cfg["slice_size"]
    overlap     = cfg["overlap"]
    target_kw   = phrase.lower().split(".")[0].strip()

    W, H  = pil_image.size
    step  = int(slice_size * (1 - overlap))
    xs    = sorted(set(list(range(0, max(1, W - slice_size + 1), step)) + [max(0, W - slice_size)]))
    ys    = sorted(set(list(range(0, max(1, H - slice_size + 1), step)) + [max(0, H - slice_size)]))
    full_prompt = " . ".join([phrase] + distractors) + " ."

    raw: list[tuple] = []
    for x0 in xs:
        for y0 in ys:
            patch = pil_image.crop((x0, y0, min(x0 + slice_size, W), min(y0 + slice_size, H)))
            inputs = _dino_processor(
                images=patch, text=full_prompt, return_tensors="pt"
            ).to(DINO_DEVICE)
            torch.cuda.empty_cache()
            with torch.no_grad():
                outputs = _dino_model(**inputs)
            res = _dino_processor.post_process_grounded_object_detection(
                outputs, inputs["input_ids"],
                threshold=threshold, text_threshold=DINO_TEXT_THR,
                target_sizes=[patch.size[::-1]],
            )[0]
            for box, score, label in zip(
                res["boxes"].cpu().numpy(),
                res["scores"].cpu().numpy(),
                res["text_labels"],
            ):
                lc = label.strip().lower()
                if target_kw in lc or lc in target_kw:
                    bx1, by1, bx2, by2 = box
                    raw.append((bx1 + x0, by1 + y0, bx2 + x0, by2 + y0, float(score)))

    if not raw:
        return {}

    # Greedy NMS (IoU > 0.5)
    raw.sort(key=lambda d: d[4], reverse=True)
    kept: list[tuple] = []
    for d in raw:
        x1, y1, x2, y2, _ = d
        ok = True
        for k in kept:
            kx1, ky1, kx2, ky2, _ = k
            ix1, iy1 = max(x1, kx1), max(y1, ky1)
            ix2, iy2 = min(x2, kx2), min(y2, ky2)
            if ix2 > ix1 and iy2 > iy1:
                inter = (ix2 - ix1) * (iy2 - iy1)
                union = (x2 - x1) * (y2 - y1) + (kx2 - kx1) * (ky2 - ky1) - inter
                if inter / union > 0.5:
                    ok = False
                    break
        if ok:
            kept.append(d)

    # Drop boxes that exceed the per-class size cap (full-image coordinates).
    max_frac = DINO_MAX_BOX_FRAC.get(cls)
    if max_frac is not None:
        W_img, H_img = pil_image.size
        img_area = W_img * H_img
        kept = [d for d in kept if (d[2] - d[0]) * (d[3] - d[1]) / img_area <= max_frac]

    return {cls: kept}


# Classes where the context gate acts as a hard blocker.
# These are the most hallucination-prone — DINO readily fires on non-clinical
# objects (baseball gloves, scarves, hats, ceiling lamps) without gating.
# All other classes run freely regardless of scene context.
DINO_GATED_CLASSES: set[str] = {"glove", "mask", "hair_net", "surgical_light"}

def _context_gate(missing_classes: list[str], yolo_dets: dict) -> list[str]:
    """
    Hard-blocks DINO_GATED_CLASSES when no context anchor is detected.
    All other classes pass through unconditionally.
    Logs context info for debugging.
    """
    gated = []
    for cls in missing_classes:
        anchors = DINO_CONTEXT_ANCHORS.get(cls)
        if cls in DINO_GATED_CLASSES and anchors:
            found = anchors.intersection(yolo_dets.keys())
            if not found:
                print(f"  [CTX] blocked {cls} — no clinical anchor detected")
                continue
            print(f"  [CTX] {cls} — anchor context: {sorted(found)}")
        elif anchors:
            found = anchors.intersection(yolo_dets.keys())
            if found:
                print(f"  [CTX] {cls} — anchor context: {sorted(found)}")
        gated.append(cls)
    return gated


def dino_infer(pil_image: Image.Image, missing_classes: list[str]) -> dict:
    """
    Run Grounding DINO for the given canonical class names.
    Only called when YOLO returned zero boxes for those classes.

    Routing:
    - DINO_SAHI classes  → _sahi_dino_query() (SAHI slicing + neg prompting)
    - DINO_ISOLATED      → one isolated _dino_query() call each
    - everything else    → one joint _dino_query() call

    Returns {canonical_class_name: [(x1, y1, x2, y2, conf), ...]}
    """
    _load_dino()

    sahi_cls = [c for c in missing_classes if c in DINO_SAHI]
    isolated = [c for c in missing_classes if c in DINO_ISOLATED and c not in DINO_SAHI]
    joint    = [c for c in missing_classes if c not in DINO_ISOLATED and c not in DINO_SAHI]
    all_dets: dict[str, list] = {}

    # ── SAHI passes (sliced + negative prompting, one class per call) ─────
    for cls in sahi_cls:
        all_dets.update(_sahi_dino_query(pil_image, cls))

    # ── Joint pass (all non-isolated, non-SAHI classes together) ─────────
    if joint:
        p2c = {DINO_FALLBACK[c][0]: c for c in joint}
        p2t = {DINO_FALLBACK[c][0]: DINO_FALLBACK[c][1] for c in joint}
        all_dets.update(_dino_query(pil_image, p2c, p2t))

    # ── Isolated passes (one class per call, no prompt competition) ───────
    for cls in isolated:
        phrase, thr = DINO_FALLBACK[cls]
        result = _dino_query(pil_image, {phrase: cls}, {phrase: thr})
        all_dets.update(result)

    return all_dets


# ══════════════════════════════════════════════════════════════════════════════
# Step 2 — Unified detection bridge  (our dict → sv.Detections for ByteTrack)
# ══════════════════════════════════════════════════════════════════════════════

def _dets_to_sv(dets_dict: dict,
                dino_classes: set[str] | None = None) -> sv.Detections | None:
    """
    Convert {class_name: [(x1,y1,x2,y2,conf),...]} into a single sv.Detections
    object using the stable CLASS_TO_ID mapping.

    Classes not in CLASS_TO_ID (e.g. future additions) fall back to id=0 so
    they still get tracked rather than silently dropped.

    dino_classes: optional set of class names that came from DINO rather than YOLO.
    Their source is stored in det.data["source"] as "dino"; all others as "yolo".
    This lets _annotate_tracked() keep DINO boxes orange after tracking.

    Returns None when dets_dict is empty.
    """
    boxes, confs, class_ids, names_list, sources = [], [], [], [], []
    for cls, det_list in dets_dict.items():
        cid = CLASS_TO_ID.get(cls, 0)
        src = "dino" if (dino_classes and cls in dino_classes) else "yolo"
        for (x1, y1, x2, y2, conf) in det_list:
            boxes.append([x1, y1, x2, y2])
            confs.append(conf)
            class_ids.append(cid)
            names_list.append(cls)
            sources.append(src)
    if not boxes:
        return None
    det = sv.Detections(
        xyxy=np.array(boxes, dtype=np.float32),
        confidence=np.array(confs, dtype=np.float32),
        class_id=np.array(class_ids, dtype=int),
    )
    det.data["class_name"] = np.array(names_list)
    det.data["source"]     = np.array(sources)
    return det


# ══════════════════════════════════════════════════════════════════════════════
# Annotation — supervision (YOLO) + cv2 overlay (DINO, orange)
# ══════════════════════════════════════════════════════════════════════════════

_DINO_BGR = (30, 100, 255)   # orange in BGR


def _build_sv_detections(dets_dict: dict):
    """
    Convert {class_name: [(x1,y1,x2,y2,conf),...]} into
    (sv.Detections, list[str] label_strings).
    """
    boxes, confs, ids, names = [], [], [], []
    for i, (cls, det_list) in enumerate(dets_dict.items()):
        for (x1, y1, x2, y2, conf) in det_list:
            boxes.append([x1, y1, x2, y2])
            confs.append(conf)
            ids.append(i)
            names.append(cls)
    if not boxes:
        return None, []
    detections = sv.Detections(
        xyxy=np.array(boxes, dtype=np.float32),
        confidence=np.array(confs, dtype=np.float32),
        class_id=np.array(ids, dtype=int),
    )
    return detections, names


def annotate_image(
    bgr: np.ndarray,
    yolo_dets: dict,
    dino_dets: dict,
) -> np.ndarray:
    """
    Draw bounding boxes and labels on a copy of the frame.
    YOLO detections use the supervision default colour palette.
    DINO detections use solid orange with a [DINO] label prefix.
    """
    scene = bgr.copy()

    # ── YOLO boxes (supervision) ───────────────────────────────────────────
    yolo_sv, yolo_names = _build_sv_detections(yolo_dets)
    if yolo_sv is not None and len(yolo_sv) > 0:
        yolo_labels = [
            f"{name}  {conf:.0%}"
            for name, conf in zip(yolo_names, yolo_sv.confidence)
        ]
        # Semi-transparent fill over each YOLO box
        overlay = scene.copy()
        for det in yolo_sv.xyxy.astype(int):
            x1b, y1b, x2b, y2b = det
            cv2.rectangle(overlay, (x1b, y1b), (x2b, y2b), (255, 255, 0), -1)
        cv2.addWeighted(overlay, 0.15, scene, 0.85, 0, scene)
        scene = sv.BoxAnnotator(thickness=2).annotate(
            scene=scene, detections=yolo_sv
        )
        scene = sv.LabelAnnotator(
            text_scale=0.5, text_thickness=1, text_padding=4
        ).annotate(scene=scene, detections=yolo_sv, labels=yolo_labels)

    # ── DINO boxes (cv2 — ensures colour independence from sv version) ─────
    if dino_dets:
        h, w = scene.shape[:2]
        font_scale = max(0.4, min(0.7, w / 1200))
        box_thick  = 2
        txt_thick  = 1
        for cls, det_list in dino_dets.items():
            for (x1, y1, x2, y2, conf) in det_list:
                x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                # Semi-transparent orange fill
                overlay = scene.copy()
                cv2.rectangle(overlay, (x1, y1), (x2, y2), _DINO_BGR, -1)
                cv2.addWeighted(overlay, 0.18, scene, 0.82, 0, scene)
                # Box: thin black outline + orange inner + white inner edge
                cv2.rectangle(scene, (x1, y1), (x2, y2), (0, 0, 0),      box_thick + 2)
                cv2.rectangle(scene, (x1, y1), (x2, y2), _DINO_BGR,      box_thick)
                cv2.rectangle(scene, (x1, y1), (x2, y2), (255, 255, 255), 1)
                # Label
                text = f"[DINO] {cls}  {conf:.0%}"
                (tw, th), bl = cv2.getTextSize(
                    text, cv2.FONT_HERSHEY_SIMPLEX, font_scale, txt_thick
                )
                pad = 5
                ty = max(y1 - pad, th + pad * 2)
                cv2.rectangle(
                    scene,
                    (x1, ty - th - pad * 2),
                    (x1 + tw + pad * 2, ty + bl),
                    (0, 0, 0), -1
                )
                cv2.rectangle(
                    scene,
                    (x1, ty - th - pad * 2),
                    (x1 + tw + pad * 2, ty + bl),
                    _DINO_BGR, 2
                )
                cv2.putText(
                    scene, text, (x1 + pad, ty - pad),
                    cv2.FONT_HERSHEY_SIMPLEX, font_scale,
                    (255, 255, 255), txt_thick, cv2.LINE_AA
                )

    return scene


# ══════════════════════════════════════════════════════════════════════════════
# Step 4a — Tracked annotation  (replaces annotate_image() in video mode)
# ══════════════════════════════════════════════════════════════════════════════

def _annotate_tracked(bgr: np.ndarray, tracked_sv: sv.Detections) -> np.ndarray:
    """
    Draw bounding boxes from ByteTrack-enriched sv.Detections.

    Each box is labelled  "<class_name> #<tracker_id>  <conf%>".
    Source separation is preserved via det.data["source"]:
      - "yolo"  →  supervision default colour palette
      - "dino"  →  solid orange (_DINO_BGR), same style as annotate_image()

    Falls back gracefully when tracker_id is None (tracker not yet warmed up
    on the very first frame, or supervision version difference).
    """
    scene = bgr.copy()
    if tracked_sv is None or len(tracked_sv) == 0:
        return scene

    class_names = tracked_sv.data.get("class_name", np.array([]))
    sources     = tracked_sv.data.get("source",     np.array(["yolo"] * len(tracked_sv)))
    tracker_ids = tracked_sv.tracker_id   # None or int array

    yolo_mask = np.array([s == "yolo" for s in sources], dtype=bool)
    dino_mask = ~yolo_mask

    # ── YOLO tracked boxes (supervision palette) ───────────────────────────
    if yolo_mask.any():
        yolo_sv = tracked_sv[yolo_mask]
        yolo_cls_names = class_names[yolo_mask]
        yolo_tids      = yolo_sv.tracker_id
        yolo_labels = []
        for i, (cls_name, conf) in enumerate(zip(yolo_cls_names, yolo_sv.confidence)):
            tid   = int(yolo_tids[i]) if yolo_tids is not None else None
            label = (f"{cls_name} #{tid}  {conf:.0%}"
                     if tid is not None else f"{cls_name}  {conf:.0%}")
            yolo_labels.append(label)
        # Semi-transparent fill
        overlay = scene.copy()
        for det in yolo_sv.xyxy.astype(int):
            x1b, y1b, x2b, y2b = det
            cv2.rectangle(overlay, (x1b, y1b), (x2b, y2b), (255, 255, 0), -1)
        cv2.addWeighted(overlay, 0.15, scene, 0.85, 0, scene)
        scene = sv.BoxAnnotator(thickness=2).annotate(scene=scene, detections=yolo_sv)
        scene = sv.LabelAnnotator(
            text_scale=0.5, text_thickness=1, text_padding=4
        ).annotate(scene=scene, detections=yolo_sv, labels=yolo_labels)

    # ── DINO tracked boxes (orange, cv2) ───────────────────────────────────
    if dino_mask.any():
        dino_sv        = tracked_sv[dino_mask]
        dino_cls_names = class_names[dino_mask]
        dino_sources   = sources[dino_mask]
        dino_tids      = dino_sv.tracker_id
        h, w           = scene.shape[:2]
        font_scale     = max(0.4, min(0.7, w / 1200))
        for i, (box, conf) in enumerate(zip(dino_sv.xyxy.astype(int), dino_sv.confidence)):
            cls_name = dino_cls_names[i]
            src_name = str(dino_sources[i])
            tid      = int(dino_tids[i]) if dino_tids is not None else None
            x1, y1, x2, y2 = box
            prefix = "[STAB]" if src_name == "stabilized" else "[DINO]"
            label = (f"{prefix} {cls_name} #{tid}  {conf:.0%}"
                     if tid is not None else f"{prefix} {cls_name}  {conf:.0%}")
            # Semi-transparent orange fill
            overlay = scene.copy()
            cv2.rectangle(overlay, (x1, y1), (x2, y2), _DINO_BGR, -1)
            cv2.addWeighted(overlay, 0.18, scene, 0.82, 0, scene)
            # Box: thin black outline + orange inner + white inner edge
            cv2.rectangle(scene, (x1, y1), (x2, y2), (0, 0, 0),      4)
            cv2.rectangle(scene, (x1, y1), (x2, y2), _DINO_BGR,      2)
            cv2.rectangle(scene, (x1, y1), (x2, y2), (255, 255, 255), 1)
            # Label background + text
            (tw, th), bl = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, font_scale, 1)
            pad = 5
            ty  = max(y1 - pad, th + pad * 2)
            cv2.rectangle(scene, (x1, ty - th - pad * 2), (x1 + tw + pad * 2, ty + bl), (0, 0, 0), -1)
            cv2.rectangle(scene, (x1, ty - th - pad * 2), (x1 + tw + pad * 2, ty + bl), _DINO_BGR, 2)
            cv2.putText(scene, label, (x1 + pad, ty - pad),
                        cv2.FONT_HERSHEY_SIMPLEX, font_scale,
                        (255, 255, 255), 1, cv2.LINE_AA)

    return scene


# ══════════════════════════════════════════════════════════════════════════════
# Step 4b — Persistent room state  (flicker-free PPE alert foundation)
# ══════════════════════════════════════════════════════════════════════════════

# PPE compliance classes and their required co-worker anchors.
# Used to detect PPE violations: a tracked worker visible for > WORKER_PPE_GRACE
# frames without a matching PPE track nearby is flagged in room_state.
_PPE_CLASSES:    frozenset[str] = frozenset({"glove", "mask", "hair_net"})
_WORKER_CLASSES: frozenset[str] = frozenset({"healthcare_worker", "person"})
# Frames a track must be continuously present before absence is logged as stale.
_ROOM_STATE_STALE_FRAMES = 150   # ~5 s at 30 fps; prune entries not seen in this window


def _box_center(box: np.ndarray) -> tuple[float, float]:
    return float((box[0] + box[2]) * 0.5), float((box[1] + box[3]) * 0.5)


def _clip_box(box: np.ndarray, h: int, w: int) -> np.ndarray:
    clipped = box.copy()
    clipped[0] = np.clip(clipped[0], 0, w - 1)
    clipped[1] = np.clip(clipped[1], 0, h - 1)
    clipped[2] = np.clip(clipped[2], 0, w - 1)
    clipped[3] = np.clip(clipped[3], 0, h - 1)
    return clipped


def _associate_worker_track(ppe_box: np.ndarray, workers: dict) -> tuple[str, np.ndarray] | tuple[None, None]:
    """Pick the nearest worker track for a PPE box (prefers containment)."""
    if not workers:
        return None, None
    cx, cy = _box_center(ppe_box)

    containing = []
    nearest = []
    for wk, wbox in workers.items():
        wx1, wy1, wx2, wy2 = wbox
        if wx1 <= cx <= wx2 and wy1 <= cy <= wy2:
            containing.append((wk, wbox))
            continue
        wcx, wcy = _box_center(wbox)
        nearest.append((float((wcx - cx) ** 2 + (wcy - cy) ** 2), wk, wbox))

    if containing:
        best = min(containing, key=lambda it: float((it[1][2] - it[1][0]) * (it[1][3] - it[1][1])))
        return best[0], best[1]

    if not nearest:
        return None, None
    nearest.sort(key=lambda it: it[0])
    return nearest[0][1], nearest[0][2]


def _stabilise_ppe_with_worker_motion(
    tracked_sv: sv.Detections,
    motion_state: dict,
    frame_idx: int,
    fps: float,
    frame_shape: tuple[int, int],
) -> sv.Detections:
    """
    Motion-aware PPE stabilizer.

    If a PPE track temporarily disappears between DINO refreshes, synthesize a
    short-lived box by translating its last known PPE box with the movement of
    the associated worker track.
    """
    if tracked_sv is None or len(tracked_sv) == 0 or tracked_sv.tracker_id is None:
        return tracked_sv

    h, w = frame_shape
    max_age = max(1, round(fps * DINO_VIDEO_INTERVAL_SEC))
    decay = 0.92
    min_conf = 0.20
    ppe_memory = motion_state.setdefault("ppe_memory", {})

    class_names = tracked_sv.data.get("class_name", np.array([""] * len(tracked_sv)))
    tracker_ids = tracked_sv.tracker_id

    workers_now: dict[str, np.ndarray] = {}
    observed_ppe_keys: set[tuple[str, int]] = set()

    for i in range(len(tracked_sv)):
        cls = str(class_names[i])
        tid = int(tracker_ids[i])
        box = tracked_sv.xyxy[i].astype(np.float32)
        if cls in _WORKER_CLASSES:
            workers_now[f"{cls}_{tid}"] = box

    for i in range(len(tracked_sv)):
        cls = str(class_names[i])
        if cls not in _PPE_CLASSES:
            continue
        tid = int(tracker_ids[i])
        key = (cls, tid)
        observed_ppe_keys.add(key)

        box = tracked_sv.xyxy[i].astype(np.float32)
        wk, wbox = _associate_worker_track(box, workers_now)
        ppe_memory[key] = {
            "box": box,
            "frame": frame_idx,
            "worker_key": wk,
            "worker_box": wbox.copy() if wbox is not None else None,
            "conf": float(tracked_sv.confidence[i]),
        }

    synth_boxes = []
    synth_confs = []
    synth_class_ids = []
    synth_names = []
    synth_sources = []
    synth_tids = []

    for key, state in list(ppe_memory.items()):
        cls, tid = key
        age = frame_idx - int(state["frame"])
        if age > max_age:
            del ppe_memory[key]
            continue
        if key in observed_ppe_keys:
            continue

        worker_key = state.get("worker_key")
        prev_worker_box = state.get("worker_box")
        if not worker_key or prev_worker_box is None or worker_key not in workers_now:
            continue

        cur_worker_box = workers_now[worker_key]
        prev_cx, prev_cy = _box_center(np.asarray(prev_worker_box, dtype=np.float32))
        cur_cx, cur_cy = _box_center(cur_worker_box)
        dx, dy = (cur_cx - prev_cx), (cur_cy - prev_cy)

        new_box = np.asarray(state["box"], dtype=np.float32).copy()
        new_box[[0, 2]] += dx
        new_box[[1, 3]] += dy
        new_box = _clip_box(new_box, h, w)
        if new_box[2] <= new_box[0] or new_box[3] <= new_box[1]:
            continue

        conf = max(float(state["conf"]) * (decay ** age), min_conf)
        synth_boxes.append(new_box)
        synth_confs.append(conf)
        synth_class_ids.append(CLASS_TO_ID.get(cls, 0))
        synth_names.append(cls)
        synth_sources.append("stabilized")
        synth_tids.append(int(tid))

        state["box"] = new_box
        state["worker_box"] = cur_worker_box.copy()

    if not synth_boxes:
        return tracked_sv

    merged = sv.Detections(
        xyxy=np.concatenate([tracked_sv.xyxy, np.asarray(synth_boxes, dtype=np.float32)], axis=0),
        confidence=np.concatenate([tracked_sv.confidence, np.asarray(synth_confs, dtype=np.float32)], axis=0),
        class_id=np.concatenate([tracked_sv.class_id, np.asarray(synth_class_ids, dtype=int)], axis=0),
    )
    merged.tracker_id = np.concatenate([tracked_sv.tracker_id, np.asarray(synth_tids, dtype=int)], axis=0)

    existing_names = tracked_sv.data.get("class_name", np.array([""] * len(tracked_sv), dtype=object))
    existing_sources = tracked_sv.data.get("source", np.array(["yolo"] * len(tracked_sv), dtype=object))
    merged.data["class_name"] = np.concatenate([existing_names, np.asarray(synth_names, dtype=object)], axis=0)
    merged.data["source"] = np.concatenate([existing_sources, np.asarray(synth_sources, dtype=object)], axis=0)
    return merged


# ══════════════════════════════════════════════════════════════════════════════
# Glare robustness — bright frame detection
# ══════════════════════════════════════════════════════════════════════════════

def _is_overexposed(bgr: np.ndarray, threshold: float = 220.0) -> bool:
    """
    Detect if a frame is severely overexposed (bright glare/light blowout).
    Returns True if mean pixel brightness exceeds threshold.
    Prevents model inference on frames the model was not trained to handle.
    """
    mean_brightness = np.mean(bgr)
    return mean_brightness > threshold


def _promote_worker_aliases(yolo_dets: dict) -> dict:
    """
    Conservative relabeling for clinical scenes.

    When YOLO returns only `person` for staff that are clearly wearing PPE, we
    promote that class to `healthcare_worker` so downstream tracking and logs
    stay semantically stable.
    """
    if "person" not in yolo_dets or "healthcare_worker" in yolo_dets:
        return yolo_dets

    clinical_context = {"mask", "glove", "hair_net"}
    if not clinical_context.intersection(yolo_dets.keys()):
        return yolo_dets

    promoted = dict(yolo_dets)
    promoted["healthcare_worker"] = promoted.pop("person")
    print("  [ALIAS] promoted person -> healthcare_worker")
    return promoted


# ══════════════════════════════════════════════════════════════════════════════
# Per-class ID stabiliser  (post-ByteTrack remapping with affinity tracking)
# ══════════════════════════════════════════════════════════════════════════════

def _stabilise_tracker_ids(
    tracked_sv: sv.Detections,
    id_remap: dict,
    frame_idx: int,
    reuse_window: int = 72,          # frames — match lost_track_buffer (3 s @ 24 fps)
) -> sv.Detections:
    """
    Map ByteTrack's ever-increasing global tracker IDs to stable per-class IDs with affinity.

    Problem: ByteTrack increments a global counter every time ANY object is lost
    and re-detected.  In a busy OR scene (monitors, beds, IV stands appearing and
    disappearing) the surgeon's mask might be displayed as #3 at frame 0 and #26
    at frame 191 despite being continuously visible — purely because other objects
    cycled IDs in between.

    Fix: maintain a per-class mapping  raw_tracker_id → stable_id WITH AFFINITY.
    When ByteTrack assigns a *new* raw_id to a class that recently had an active
    track (within `reuse_window` frames), we prefer to recycle the stable_id that
    that raw_id historically held (affinity), or else the most-recently-seen orphaned
    stable_id.  This prevents ID collapse during glare occlusions: if hair_net's
    raw_id=7 held stable_id=#1, and it re-appears after glare as raw_id=9, it
    still gets stable_id=#1 because we remember the affinity.

    For multi-instance classes (two monitors in frame) each instance keeps its own
    stable ID (1 and 2), and if one disappears and reappears it recycles back.

    id_remap (mutated in-place between calls):
        { class_name: {
            "raw_to_stable": { raw_id (int): stable_id (int), ... },
            "stable_to_raw_affinity": { stable_id (int): raw_id (int), ... },
            "counter":       int,   # next stable_id to assign
            "last_seen":     { stable_id (int): frame_idx (int) },
          }, ...
        }

    Returns a shallow-copied sv.Detections with tracker_id replaced by stable IDs.
    """
    import copy as _copy
    from collections import defaultdict as _dd

    if tracked_sv is None or len(tracked_sv) == 0 or tracked_sv.tracker_id is None:
        return tracked_sv

    class_names = tracked_sv.data.get("class_name", np.array([]))
    raw_ids     = tracked_sv.tracker_id.copy()
    stable_ids  = raw_ids.copy()

    # Group indices by class so we can reason about all instances together
    by_class: dict[str, list[tuple[int, int]]] = _dd(list)
    for i, (cls, raw_id) in enumerate(zip(class_names, raw_ids)):
        by_class[str(cls)].append((i, int(raw_id)))

    for cls, items in by_class.items():
        if cls not in id_remap:
            id_remap[cls] = {
                "raw_to_stable": {},
                "stable_to_raw_affinity": {},  # Track affinity: stable_id → last raw_id that held it
                "counter": 0,
                "last_seen": {}
            }
        state = id_remap[cls]
        r2s   = state["raw_to_stable"]
        affinity = state["stable_to_raw_affinity"]

        for (i, raw_id) in items:
            if raw_id in r2s:
                # Already known — keep existing stable ID and refresh timestamp
                sid = r2s[raw_id]
                state["last_seen"][sid] = frame_idx
                affinity[sid] = raw_id  # Update affinity record
                stable_ids[i] = sid
            else:
                # New raw ID from ByteTrack.
                # Look for an orphaned stable ID: one that was recently active
                # for this class but currently has no raw_id pointing at it.
                current_stable = {r2s[r] for (_, r) in items if r in r2s}
                reusable = [
                    sid for sid, last_f in state["last_seen"].items()
                    if sid not in current_stable
                    and (frame_idx - last_f) <= reuse_window
                ]
                if reusable:
                    # Recycle the most recently seen orphaned stable ID
                    sid = max(reusable, key=lambda s: state["last_seen"][s])
                else:
                    # Genuinely new object (or too long since last seen)
                    state["counter"] += 1
                    sid = state["counter"]
                r2s[raw_id]          = sid
                affinity[sid]        = raw_id  # Record affinity for future re-entry
                state["last_seen"][sid] = frame_idx
                stable_ids[i]        = sid

    stabilised            = _copy.copy(tracked_sv)
    stabilised.tracker_id = stable_ids
    return stabilised


def _update_room_state(room_state: dict, tracked_sv: sv.Detections,
                       frame_idx: int) -> None:
    """
    Upsert per-track state into room_state.

    Key  : "{class_name}_{tracker_id}"  (e.g. "healthcare_worker_3")
    Value: {
        "class":      str,
        "tracker_id": int,
        "first_seen": int,   # frame index when this track was first created
        "last_seen":  int,   # most recent frame index
        "conf":       float, # latest confidence
        "box":        [x1, y1, x2, y2],
    }

    Stale entries (not seen for > _ROOM_STATE_STALE_FRAMES frames) are pruned
    every 30 frames to prevent unbounded growth during long videos.
    """
    if tracked_sv is not None and len(tracked_sv) > 0:
        class_names = tracked_sv.data.get("class_name", np.array([]))
        tracker_ids = tracked_sv.tracker_id
        if tracker_ids is not None:
            for i, tid in enumerate(tracker_ids):
                if tid is None:
                    continue
                cls = str(class_names[i])
                key = f"{cls}_{int(tid)}"
                existing = room_state.get(key)
                room_state[key] = {
                    "class":      cls,
                    "tracker_id": int(tid),
                    "first_seen": existing["first_seen"] if existing else frame_idx,
                    "last_seen":  frame_idx,
                    "conf":       float(tracked_sv.confidence[i]),
                    "box":        tracked_sv.xyxy[i].tolist(),
                }

    # Prune stale tracks every 30 frames
    if frame_idx % 30 == 0:
        cutoff = frame_idx - _ROOM_STATE_STALE_FRAMES
        stale  = [k for k, v in room_state.items() if v["last_seen"] < cutoff]
        for k in stale:
            del room_state[k]


# ══════════════════════════════════════════════════════════════════════════════
# Excel helpers  (same schema as infer_v3.py / infer_ensemble.py)
# ══════════════════════════════════════════════════════════════════════════════

def get_or_create_workbook():
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        col_map = {}
        for cell in ws[1]:
            if cell.column > NUM_FIXED and cell.value:
                col_map[str(cell.value)] = cell.column
        return wb, ws, col_map

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "HospitalGuard-109 Log"
    for ci, h in enumerate(FIXED_HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL_FIXED
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    for col, w in FIXED_WIDTHS.items():
        ws.column_dimensions[col].width = w
    return wb, ws, {}


def _ensure_class_col(ws, col_map: dict, class_name: str) -> int:
    if class_name not in col_map:
        new_col = NUM_FIXED + len(col_map) + 1
        col_map[class_name] = new_col
        cell = ws.cell(row=1, column=new_col, value=class_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL_CLASS
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(new_col)].width = 18
    return col_map[class_name]


def write_row(ws, col_map, row_idx, url, target,
              detected_str, conf_str, result_type, notes, flat_dets):
    ws.cell(row=row_idx, column=1, value=url)
    ws.cell(row=row_idx, column=2, value=target)
    ws.cell(row=row_idx, column=3, value=detected_str)
    ws.cell(row=row_idx, column=4, value=conf_str)
    ws.cell(row=row_idx, column=5, value=result_type)
    ws.cell(row=row_idx, column=6, value=notes or "")
    for cls_name, confs in flat_dets.items():
        col_idx = _ensure_class_col(ws, col_map, cls_name)
        ws.cell(
            row=row_idx, column=col_idx,
            value=", ".join(f"{c:.2f}" for c in sorted(confs, reverse=True))
        )
    color = RESULT_COLORS.get(result_type, "FFFFFF")
    fill  = PatternFill("solid", fgColor=color)
    align = Alignment(wrap_text=True, vertical="top")
    for col in range(1, NUM_FIXED + 1):
        cell            = ws.cell(row=row_idx, column=col)
        cell.fill       = fill
        cell.alignment  = align


def log_entry(url: str, target: str, flat_dets: dict, extra_notes: str = "") -> str:
    """
    Append one row to the Excel log and save atomically.
    Called automatically by run_image() and the video branch — no manual
    logging needed regardless of how inference is triggered.
    Returns result_type string.
    """
    detected_str, conf_str, result_type, auto_notes = classify_result(target, flat_dets)
    notes = " | ".join(filter(None, [extra_notes, auto_notes]))
    wb, ws, col_map = get_or_create_workbook()
    write_row(ws, col_map, ws.max_row + 1,
              url, target, detected_str, conf_str, result_type, notes, flat_dets)
    wb.save(EXCEL_PATH)
    return result_type


# ══════════════════════════════════════════════════════════════════════════════
# Result classification  (identical to infer_ensemble.py)
# ══════════════════════════════════════════════════════════════════════════════

def classify_result(target: str, flat_dets: dict):
    """
    flat_dets = {cls: [conf, ...]}
    Returns (detected_str, conf_str, result_type, notes).
    """
    t_lower = target.strip().lower()
    is_none = t_lower in {"[none]", "none", ""}

    if is_none:
        if flat_dets:
            fp = ", ".join(
                f"{k} ({max(v):.2f})" for k, v in list(flat_dets.items())[:8]
            )
            return "Yes", "N/A", "FP (False Positive)", f"Model detected: {fp}"
        return "No", "N/A", "TN (True Negative)", "Nothing detected as expected."

    matched = next((k for k in flat_dets if k.lower() == t_lower), None)
    if matched:
        confs    = sorted(flat_dets[matched], reverse=True)
        conf_str = ", ".join(f"{c:.2f}" for c in confs)
        rtype    = "TP" if confs[0] >= LOW_CONF_THRESH else "TP (Low Conf)"
        others   = {k: v for k, v in flat_dets.items() if k != matched}
        notes    = (
            "Also detected: "
            + ", ".join(f"{k} ({max(v):.2f})" for k, v in list(others.items())[:8])
        ) if others else ""
        return "Yes", conf_str, rtype, notes

    if flat_dets:
        alt = ", ".join(
            f"{k} ({max(v):.2f})" for k, v in list(flat_dets.items())[:5]
        )
        notes = f"Model saw instead: {alt}"
    else:
        notes = "Nothing detected at all."
    return "No", "N/A", "FN (False Negative)", notes


# ══════════════════════════════════════════════════════════════════════════════
# Video inference
# ══════════════════════════════════════════════════════════════════════════════

def run_video(v1: YOLO, v3: YOLO, video_path: Path, out_path: Path) -> dict:
    """
    Process a video file frame-by-frame.
    Returns a flat summary: {class_name: [conf, ...]} across all frames.

    Strategy:
      - YOLO ensemble runs on every frame.
      - Grounding DINO runs every DINO_VIDEO_INTERVAL frames, only for
        weak-class names that YOLO found zero boxes for in that frame.
    """
    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        raise RuntimeError(f"Cannot open video: {video_path}")

    fps    = cap.get(cv2.CAP_PROP_FPS) or 25.0
    width  = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    total  = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    # ── Step 1: ByteTrack initialisation ─────────────────────────────────
    # lost_track_buffer = 3 s of video at source FPS
    #   e.g. 24 fps → 72 frames of occlusion memory before an ID is dropped.
    #   Increased from 1 s: surgical occlusions (hand/scalpel covering a mask
    #   or glove) routinely last 1–2 s, which was causing ID resets on re-entry.
    #   3 s ensures PPE tracks survive typical procedural occlusions.
    # track_activation_threshold kept low (0.25) so DINO detections
    #   (which often score 0.30–0.45) can initiate tracks.
    # minimum_matching_threshold = 0.6 IoU: relaxed from 0.8 to tolerate larger
    #   box-shift between frames (surgeon turning, camera pan, partial occlusion).
    #   0.8 was too strict — a >20% shift (e.g. head turning sideways) caused
    #   ByteTrack to mint a new ID rather than re-use the existing track.
    byte_tracker = sv.ByteTrack(
        track_activation_threshold=0.25,
        lost_track_buffer=max(1, round(fps * 3)),   # 3 seconds of occlusion memory
        minimum_matching_threshold=0.6,
        frame_rate=int(fps),
    )

    writer = cv2.VideoWriter(
        str(out_path),
        cv2.VideoWriter_fourcc(*"mp4v"),
        fps,
        (width, height),
    )

    all_confs: dict[str, list] = defaultdict(list)   # for Excel summary
    room_state: dict[str, dict] = {}                 # keyed by "{class_name}_{tracker_id}"
    id_remap:   dict[str, dict] = {}                 # per-class stable ID remapping state
    motion_state: dict = {}                          # worker-anchored PPE motion stabilizer state
    last_tracked_sv: sv.Detections | None = None     # carry-forward detections on glare frames
    frame_idx = 0
    dino_frame_interval = max(1, int(DINO_VIDEO_INTERVAL_FRAMES))
    print(f"  Video: {width}x{height} @ {fps:.1f} fps  ({total} frames)")
    print(f"  DINO fires every {dino_frame_interval} frames (~{dino_frame_interval / max(fps, 1e-6):.2f}s) for missed weak classes.")

    while True:
        ret, bgr = cap.read()
        if not ret:
            break
        frame_idx += 1

        if frame_idx % 50 == 0:
            print(f"  Frame {frame_idx}/{total} …")

        # Glare gate: detect overexposed frames and skip YOLO/DINO inference.
        # Instead, carry forward the last good tracked detections to maintain
        # continuity.  Overexposed frames are brief (1-3 frames) so the carry-
        # forward is visually indistinguishable and prevents PPE disappearance.
        is_glare = _is_overexposed(bgr, threshold=220.0)

        if is_glare and last_tracked_sv is not None:
            # Glare detected and we have prior detections — reuse them.
            # This prevents ByteTrack from seeing (empty detections) and orphaning
            # tracks, which would cause ID resets on recovery.
            tracked_sv = last_tracked_sv
        else:
            # Normal frame: run YOLO + DINO
            # Layer 1: YOLO on a temp file (ultralytics needs a path or numpy array)
            yolo_dets = _yolo_on_frame(v1, v3, bgr)
            yolo_dets = _promote_worker_aliases(yolo_dets)

            # Layer 2: DINO every N frames. We intentionally do not carry
            # boxes forward here, because stale DINO boxes freeze in place
            # while the object or camera is still moving.
            active_dino: dict = {}
            if (frame_idx - 1) % dino_frame_interval == 0:
                detected_cls = set(yolo_dets.keys())
                all_dino_targets = set(DINO_FALLBACK) | set(DINO_SAHI)
                missing_weak = [c for c in all_dino_targets if c not in detected_cls]
                if missing_weak:
                    pil_img     = Image.fromarray(cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB))
                    active_dino = dino_infer(pil_img, missing_weak)

            all_dets = {**active_dino, **yolo_dets}

            # Step 2: merge into unified sv.Detections for the tracker
            # Pass active_dino keys so each detection is tagged "yolo" or "dino"
            # in det.data["source"] — preserved through tracking for correct colouring.
            combined_sv = _dets_to_sv(all_dets, dino_classes=set(active_dino.keys()))

            # Step 3: feed combined detections through ByteTrack
            # update_with_detections() returns sv.Detections enriched with tracker_id.
            # Passing empty Detections when nothing was found lets the tracker age out
            # lost tracks correctly rather than skipping the update entirely.
            tracked_sv = byte_tracker.update_with_detections(
                combined_sv if combined_sv is not None else sv.Detections.empty()
            )

            # Step 3b: remap raw ByteTrack IDs → stable per-class IDs with affinity.
            # ByteTrack's global counter increments for every lost+re-detected object
            # across ALL classes (beds, monitors, IV stands, surgeon, PPE).  A mask
            # that disappears for 2 frames returns as raw_id=26 instead of raw_id=3.
            # _stabilise_tracker_ids() maps it back to stable_id=1 for the duration
            # of the video — making labels human-readable and PPE logic consistent.
            # Affinity tracking ensures that if a raw_id re-appears after glare, it
            # reclaims its historical stable_id rather than getting a new one.
            reuse_w = max(1, round(fps * 3))   # recycle within 3 s (= lost_track_buffer)
            tracked_sv = _stabilise_tracker_ids(tracked_sv, id_remap, frame_idx, reuse_w)
            tracked_sv = _stabilise_ppe_with_worker_motion(
                tracked_sv,
                motion_state,
                frame_idx,
                fps,
                bgr.shape[:2],
            )

            # Accumulate confidences for summary
            for cls, dets in all_dets.items():
                for det in dets:
                    all_confs[cls].append(det[4])
        
        # Remember this tracked result for potential glare frame carry-forward
        last_tracked_sv = tracked_sv

        # Step 4: update persistent room state and annotate with track IDs
        # _update_room_state() upserts each tracked box into room_state keyed by
        # "{class_name}_{tracker_id}" — enables flicker-free PPE alert logic.
        # _annotate_tracked() draws labelled boxes; DINO boxes remain orange.
        _update_room_state(room_state, tracked_sv, frame_idx)
        annotated = _annotate_tracked(bgr, tracked_sv)
        writer.write(annotated)

    cap.release()
    writer.release()
    print(f"  Processed {frame_idx} frames.")
    return dict(all_confs)


def _yolo_on_frame(v1: YOLO, v3: YOLO, bgr: np.ndarray) -> dict:
    """
    Run the V1+V3 ensemble on a single BGR numpy frame.
    Ultralytics accepts numpy arrays directly — no temp file needed.
    """
    r1 = v1(bgr, conf=CONF, iou=IOU, verbose=False)[0]
    r3 = v3(bgr, conf=CONF, iou=IOU, verbose=False)[0]

    v1_dets: dict[str, list] = defaultdict(list)
    if r1.boxes is not None:
        for box in r1.boxes:
            name = _canonical_class_name(v1.names[int(box.cls)])
            xyxy = box.xyxy[0].cpu().tolist()
            v1_dets[name].append((*xyxy, float(box.conf)))

    v3_dets: dict[str, list] = defaultdict(list)
    if r3.boxes is not None:
        for box in r3.boxes:
            name = _canonical_class_name(v3.names[int(box.cls)])
            xyxy = box.xyxy[0].cpu().tolist()
            v3_dets[name].append((*xyxy, float(box.conf)))

    merged: dict[str, list] = {}
    for cls in V3_ONLY_NEW:
        if cls in v3_dets:
            merged[cls] = v3_dets[cls]
    for cls in V3_WORKING_OVERLAP:
        combined = v1_dets.get(cls, []) + v3_dets.get(cls, [])
        if combined:
            merged[cls] = _nms_merge(combined, IOU)
    exclude = V3_WORKING_OVERLAP | V3_ONLY_NEW
    for cls, dets in v1_dets.items():
        if cls not in exclude:
            merged[cls] = dets
    return _post_filter_yolo(merged)


# ══════════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════════

def _download(url: str) -> Path:
    suffix = Path(url.split("?")[0]).suffix or ".jpg"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    r.raise_for_status()
    tmp.write(r.content)
    tmp.close()
    return Path(tmp.name)


def _source_tag(cls: str, dino_dets: dict) -> str:
    if cls in dino_dets:
        return "[DINO]"
    if cls in V3_ONLY_NEW:
        return "[V3]"
    if cls in V3_WORKING_OVERLAP:
        return "[V1+V3]"
    return "[V1]"


def _is_video(url: str) -> bool:
    suffix = Path(url.split("?")[0]).suffix.lower()
    return suffix in VIDEO_EXTS


def _is_image(url: str) -> bool:
    suffix = Path(url.split("?")[0]).suffix.lower()
    # Treat unknown extensions as images (most URL image links omit extension)
    return suffix in IMAGE_EXTS or suffix not in VIDEO_EXTS


# ══════════════════════════════════════════════════════════════════════════════
# Main loop
# ══════════════════════════════════════════════════════════════════════════════

def run_image(v1: YOLO, v3: YOLO, media_path: Path,
              url: str = "", target: str = "[None]") -> tuple[dict, Path]:
    """
    Full image inference pipeline: YOLO ensemble → DINO fallback → annotate → save → log.
    Always logs to EXCEL_PATH via log_entry(), regardless of the caller.
    Returns (all_dets, out_path).
    """
    # ── Layer 1: YOLO ensemble ─────────────────────────────────────────────
    yolo_dets = ensemble_infer(v1, v3, media_path)

    # ── Layer 1b: YOLO SAHI for small-object classes ──────────────────────
    # Runs YOLO on overlapping 640-px crops so tiny objects (fire extinguisher
    # on a wall, scissors on a tray, IV stand pole) are large enough to detect.
    sahi_supplement = _yolo_sahi_supplement(v1, v3, media_path)
    for cls, dets in sahi_supplement.items():
        if cls in yolo_dets:
            yolo_dets[cls] = _nms_merge(yolo_dets[cls] + dets, IOU)
        else:
            yolo_dets[cls] = dets

    yolo_dets = _promote_worker_aliases(yolo_dets)

    # ── Layer 1c: Surgical-scene bottle suppression ────────────────────────
    if "bottle" in yolo_dets:
        suppressor_found = BOTTLE_SURGICAL_SUPPRESSOR.intersection(yolo_dets.keys())
        if len(suppressor_found) >= BOTTLE_SURGICAL_QUORUM:
            del yolo_dets["bottle"]
            print(f"  [CTX] suppressed bottle — surgical scene ({sorted(suppressor_found)})")

    # ── Layer 2: DINO for missing weak classes ─────────────────────────────
    all_dino_targets = set(DINO_FALLBACK) | set(DINO_SAHI)
    missing_weak = [c for c in all_dino_targets if c not in yolo_dets]
    missing_weak = _context_gate(missing_weak, yolo_dets)
    dino_dets: dict = {}
    if missing_weak:
        pil_img = Image.open(media_path).convert("RGB")
        dino_dets = dino_infer(pil_img, missing_weak)
        if dino_dets:
            print(f"  [DINO] filled in: {sorted(dino_dets.keys())}")

    # ── Layer 2b: hair_net proximity filter ───────────────────────────────
    # Keep only hair_net detections whose centre falls STRICTLY INSIDE a
    # detected person / healthcare_worker bounding box.
    # No margin — hair nets are worn on heads, so the box centre must
    # overlap the worker's body box directly.
    if "hair_net" in dino_dets:
        person_boxes = [
            d[:4] for cls in ("person", "healthcare_worker")
            for d in yolo_dets.get(cls, [])
        ]
        def _inside_person(box, pboxes):
            cx = (box[0] + box[2]) / 2
            cy = (box[1] + box[3]) / 2
            for px1, py1, px2, py2, *_ in pboxes:
                if px1 <= cx <= px2 and py1 <= cy <= py2:
                    return True
            return False
        kept = [d for d in dino_dets["hair_net"] if _inside_person(d, person_boxes)]
        if len(kept) < len(dino_dets["hair_net"]):
            dropped = len(dino_dets["hair_net"]) - len(kept)
            print(f"  [PROX] hair_net: dropped {dropped} box(es) not inside any person/worker box")
        if kept:
            dino_dets["hair_net"] = kept
        else:
            del dino_dets["hair_net"]

    all_dets  = {**yolo_dets, **dino_dets}
    flat_dets = {cls: [d[4] for d in dets] for cls, dets in all_dets.items()}

    print(f"  Detections ({len(all_dets)} classes):")
    for cls in sorted(all_dets):
        confs = [round(d[4], 3) for d in all_dets[cls]]
        tag   = _source_tag(cls, dino_dets)
        print(f"    {tag:<8} {cls}: {confs}")

    # ── Annotate + save ────────────────────────────────────────────────────
    ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem      = media_path.stem[:35] or "media"
    bgr       = cv2.imread(str(media_path))
    annotated = annotate_image(bgr, yolo_dets, dino_dets)
    out_path  = OUT_DIR / f"{stem}_{OUTPUT_RUN_TAG}_{ts}.jpg"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    cv2.imwrite(str(out_path), annotated)
    print(f"  Saved : {out_path.name}")

    # ── Log (always, regardless of caller) ────────────────────────────────
    extra = ""
    t_lower = target.strip().lower()
    if t_lower not in {"[none]", "none", ""}:
        matched_key = next((k for k in all_dets if k.lower() == t_lower), None)
        if matched_key:
            extra = _source_tag(matched_key, dino_dets)
    result_type = log_entry(url or str(media_path), target, flat_dets, extra)
    print(f"  Result: {result_type}")

    return all_dets, out_path


def main():
    if not torch.cuda.is_available():
        raise RuntimeError("CUDA GPU is required. This pipeline is configured for GPU-only YOLO and DINO.")

    print("Loading V1 (106-class hospital)...")
    v1 = YOLO(str(V1_PATH))
    print("Loading V3 (109-class)...")
    v3 = YOLO(str(V3_PATH))
    v1.to("cuda")
    v3.to("cuda")

    print("\nHospitalGuard-109 ready.")
    print(f"  V1+V3 NMS overlap : {sorted(V3_WORKING_OVERLAP)}")
    print(f"  V3-only classes   : {sorted(V3_ONLY_NEW)}")
    print(f"  DINO fallback     : {sorted(DINO_FALLBACK.keys())}  (AP50 < 0.25)")
    print(f"  DINO video rate   : every {DINO_VIDEO_INTERVAL_FRAMES} frames")
    print(f"  Output dir        : {OUT_DIR.relative_to(ROOT_DIR)}")
    print(f"  Excel log         : {EXCEL_PATH.relative_to(ROOT_DIR)}")
    print(f"\nPaste an image or video URL (or 'quit' to exit).\n")

    while True:
        url = input("URL: ").strip()
        if url.lower() in {"quit", "exit", "q", "stop"}:
            break
        if not url:
            continue

        target = input("Expected class (or [None]): ").strip() or "[None]"

        media_path: Path | None = None
        is_temp    = False
        try:
            # Accept local file paths as well as remote URLs
            local = Path(url)
            if local.exists() and local.is_file():
                media_path = local
                print(f"  Local file: {local}")
            else:
                media_path = _download(url)
                is_temp    = True
                print("  Downloaded.")

            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            stem = media_path.stem[:35] or "media"

            # ── Route: video vs image ──────────────────────────────────────
            if _is_video(str(media_path)):
                print(f"  Mode: VIDEO  (DINO every {DINO_VIDEO_INTERVAL_SEC}s)")
                out_path  = OUT_DIR / f"{stem}_{OUTPUT_RUN_TAG}_{ts}.mp4"
                all_confs = run_video(v1, v3, media_path, out_path)
                print(f"  Saved : {out_path.name}")

                flat_dets = {
                    cls: [max(confs)] for cls, confs in all_confs.items()
                }
                print(f"  Classes seen across video ({len(flat_dets)}):")
                for cls in sorted(flat_dets, key=lambda c: flat_dets[c][0], reverse=True):
                    print(f"    {cls}: max_conf={flat_dets[cls][0]:.3f}  "
                          f"detections={len(all_confs[cls])}")

                _, conf_str, result_type, _ = classify_result(target, flat_dets)
                log_entry(url, target, flat_dets, f"[VIDEO {stem}]")
                print(f"  Result: {result_type}  |  conf: {conf_str}")

            else:
                print("  Mode: IMAGE")
                run_image(v1, v3, media_path, url, target)

        except Exception as e:
            print(f"  Error: {e}")
        finally:
            if is_temp and media_path and media_path.exists():
                os.unlink(media_path)

    print(f"\nDone. Log saved → {EXCEL_PATH}")


if __name__ == "__main__":
    main()
