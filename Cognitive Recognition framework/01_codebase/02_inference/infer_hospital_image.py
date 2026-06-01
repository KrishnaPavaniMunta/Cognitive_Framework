"""
infer_hospital_image.py
Interactive inference loop using the trained 106-class hospital model.
Run it, then paste image paths or URLs one at a time. Results are saved
to hospital_inference.xlsx after every image.

    python infer_hospital_image.py

Press Ctrl+C or type quit / exit to stop.
"""

import cv2
import os
import tempfile
from datetime import datetime
from pathlib import Path

import requests
from ultralytics import YOLO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR   = Path(__file__).parent.resolve()
MODEL_PATH = BASE_DIR / "runs" / "hospital" / "phase1_head" / "weights" / "best.pt"
EXCEL_PATH = BASE_DIR / "hospital_inference.xlsx"
OUTPUT_DIR = BASE_DIR / "hospital_results"
MODEL_NAME = "yolo26m-hospital-phase1"
CONF       = 0.25
IOU        = 0.45

HEADERS = [
    "Timestamp",
    "Image_File",
    "Target_Class",
    "Model",
    "Lighting",
    "Viewpoint",
    "Distance",
    "Detected",
    "Num_Instances",
    "Max_Confidence",
    "Avg_Confidence",
    "All_Detections",
    "Total_Objects",
    "Annotated_Image",
    "Notes",
]

STOP_WORDS = {"quit", "exit", "q", "stop"}


def get_or_create_workbook():
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hospital Inference"
        ws.append(HEADERS)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        col_widths = {
            "A": 20, "B": 32, "C": 20, "D": 30, "E": 18,
            "F": 15, "G": 15, "H": 10, "I": 14, "J": 16,
            "K": 16, "L": 55, "M": 14, "N": 45, "O": 30,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
    return wb, ws


def download_image(url: str) -> Path:
    print("  Downloading...")
    hdrs = {"User-Agent": "Mozilla/5.0"}
    with requests.get(url, headers=hdrs, stream=True, timeout=30) as r:
        r.raise_for_status()
        ct  = r.headers.get("Content-Type", "")
        ext = ".png" if "png" in ct else ".webp" if "webp" in ct else ".jpg"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        for chunk in r.iter_content(65536):
            tmp.write(chunk)
        tmp.close()
    return Path(tmp.name)


def infer_one(model, class_names, source: str, target: str, lighting: str):
    is_url   = source.startswith("http://") or source.startswith("https://")
    tmp_path = None

    if is_url:
        tmp_path     = download_image(source)
        img_path     = tmp_path
        display_name = source.split("/")[-1].split("?")[0] or "url_image.jpg"
    else:
        img_path = Path(source)
        if not img_path.exists():
            print(f"  ERROR: File not found -> {source}")
            return
        display_name = img_path.name

    name_to_idx  = {v.lower(): k for k, v in class_names.items()}
    target_lower = target.strip().lower() if target.strip() else None

    if target_lower:
        if target_lower not in name_to_idx:
            hints    = [n for n in name_to_idx if target_lower in n]
            hint_str = f"  Did you mean: {hints}" if hints else ""
            print(f"  WARNING: '{target}' not in model classes. Running without filter.{hint_str}")
            target_lower = None

    target_idx   = name_to_idx[target_lower] if target_lower else None
    target_label = target.strip() if target_lower else "all"

    results = model.predict(source=str(img_path), conf=CONF, iou=IOU, save=False, verbose=False)
    result  = results[0]
    boxes   = result.boxes

    total_obj = len(boxes)
    det_parts = [f"{class_names[int(b.cls[0])]}:{float(b.conf[0]):.2f}" for b in boxes]
    all_dets  = ", ".join(det_parts) if det_parts else "none"

    target_boxes  = [b for b in boxes if int(b.cls[0]) == target_idx] if target_idx is not None else list(boxes)
    num_instances = len(target_boxes)
    detected      = "Yes" if num_instances > 0 else "No"

    if num_instances > 0:
        confs    = [float(b.conf[0]) for b in target_boxes]
        max_conf = round(max(confs), 4)
        avg_conf = round(sum(confs) / len(confs), 4)
    else:
        max_conf = avg_conf = 0.0

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stem           = Path(display_name).stem
    suffix         = Path(display_name).suffix or ".jpg"
    annotated_path = OUTPUT_DIR / f"{stem}_annotated{suffix}"
    cv2.imwrite(str(annotated_path), result.plot())

    sep = "  " + chr(9472)*46
    print(f"\n{sep}")
    print(f"  Target    : {target_label}")
    print(f"  Detected  : {detected}  ({num_instances} instance(s))")
    if num_instances > 0:
        print(f"  Max conf  : {max_conf:.4f}  |  Avg conf: {avg_conf:.4f}")
    print(f"  All boxes : {all_dets}")
    print(f"  Saved to  : {annotated_path.name}")
    print(f"{sep}\n")

    wb, ws = get_or_create_workbook()
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        display_name,
        target_label,
        MODEL_NAME,
        lighting if lighting else "",
        "",
        "",
        detected,
        num_instances,
        max_conf,
        avg_conf,
        all_dets,
        total_obj,
        str(annotated_path),
        "",
    ])
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    det_cell = ws.cell(row=last_row, column=8)
    if detected == "Yes":
        det_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        det_cell.font = Font(color="276221", bold=True)
    else:
        det_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        det_cell.font = Font(color="9C0006", bold=True)
    wb.save(EXCEL_PATH)
    print(f"  Saved to Excel -> {EXCEL_PATH.name}")

    if tmp_path:
        os.unlink(tmp_path)


def main():
    print("=" * 54)
    print("  Hospital Inference  --  Interactive Mode")
    print(f"  Model : {MODEL_PATH.name}")
    print(f"  Excel : {EXCEL_PATH.name}")
    print(f"  Conf  : {CONF}  |  IoU: {IOU}")
    print("  Type 'quit' or press Ctrl+C to stop")
    print("=" * 54)

    print("\nLoading model...")
    model       = YOLO(str(MODEL_PATH))
    class_names = model.names
    print(f"Model loaded -- {len(class_names)} classes\n")

    count = 0
    while True:
        try:
            print(f"--- Image #{count + 1} " + "-" * 36)
            source = input("  Image (path or URL) : ").strip()
            if not source or source.lower() in STOP_WORDS:
                break

            target   = input("  Target class (Enter = all) : ").strip()
            if target.lower() in STOP_WORDS:
                break

            lighting = input("  Lighting condition         : ").strip()

            infer_one(model, class_names, source, target, lighting)
            count += 1

        except KeyboardInterrupt:
            print("\n\nStopped by user.")
            break
        except Exception as e:
            print(f"  ERROR: {e}\n")
            continue

    print(f"\nDone. Processed {count} image(s). Results in {EXCEL_PATH.name}")


if __name__ == "__main__":
    main()
