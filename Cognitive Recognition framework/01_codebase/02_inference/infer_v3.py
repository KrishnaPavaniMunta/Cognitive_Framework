"""
infer_v3.py
───────────
Interactive inference using the V3 hospital model (109 classes, Phase 2).

Accepts image or video URLs one at a time, runs detection, and appends one
row per test to outputs/v3_inference_log.xlsx.

Excel columns
─────────────
  A  Image ID               — the URL you pasted
  B  Ground Truth (Target)  — class you expect (or [None])
  C  Detected?              — Yes / No
  D  Confidence Score       — score(s) for the target class
  E  Result Type            — TP / TP (Low Conf) / FN / FP / TN
  F  Notes                  — what else was seen / what was missed
  G+ [class name]           — one column per detected class; cell = confidence(s)

Usage:
    python infer_v3.py
    (then paste URLs one at a time; type quit to exit)
"""

import os
import sys
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import cv2
import requests
from ultralytics import YOLO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.resolve()
MODEL_PATH = (BASE_DIR / "outputs" / "runs" / "hospital_v3"
              / "phase2_neck_head" / "weights" / "best.pt")
EXCEL_PATH = BASE_DIR / "outputs" / "v3_inference_log.xlsx"
OUTPUT_DIR = BASE_DIR / "outputs" / "inference_output"

# ── Settings ───────────────────────────────────────────────────────────────────
CONF            = 0.25   # detection confidence threshold
IOU             = 0.45
LOW_CONF_THRESH = 0.50   # below this → TP (Low Conf)
VIDEO_FPS       = 1      # frames per second to sample from video

# ── Media type helpers ─────────────────────────────────────────────────────────
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff", ".gif"}
VIDEO_EXTS = {".mp4", ".avi", ".mov", ".mkv", ".webm", ".m4v"}
STOP_WORDS = {"quit", "exit", "q", "stop"}

# ── Fixed Excel columns ────────────────────────────────────────────────────────
FIXED_HEADERS = [
    "Image ID",
    "Ground Truth (Target)",
    "Detected?",
    "Confidence Score",
    "Result Type",
    "Notes (False Positives/Misses)",
]
NUM_FIXED = len(FIXED_HEADERS)

FIXED_WIDTHS = {"A": 55, "B": 24, "C": 12, "D": 22, "E": 22, "F": 60}

# ── Result colour coding ───────────────────────────────────────────────────────
RESULT_COLORS = {
    "TP":                  "C6EFCE",   # green
    "TP (Low Conf)":       "FFEB9C",   # yellow
    "FN (False Negative)": "FFC7CE",   # red
    "FP (False Positive)": "FFCC99",   # orange
    "TN (True Negative)":  "D9E1F2",   # blue-grey
}

HEADER_FILL_FIXED = PatternFill("solid", fgColor="1D1E37")   # navy
HEADER_FILL_CLASS = PatternFill("solid", fgColor="FA643F")   # orange
HEADER_FONT       = Font(bold=True, color="FFFFFF")


# ── Workbook helpers ───────────────────────────────────────────────────────────

def get_or_create_workbook():
    """
    Open existing workbook or create a new one.
    Returns (wb, ws, col_map) where col_map = {class_name: col_index (1-based)}.
    """
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        col_map = {}
        for cell in ws[1]:
            if cell.column > NUM_FIXED and cell.value:
                col_map[str(cell.value)] = cell.column
        return wb, ws, col_map

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "V3 Inference Log"

    for ci, h in enumerate(FIXED_HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_FIXED
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    ws.row_dimensions[1].height = 32
    for col, w in FIXED_WIDTHS.items():
        ws.column_dimensions[col].width = w

    return wb, ws, {}


def ensure_class_column(ws, col_map: dict, class_name: str) -> int:
    """Add a class column to the header row if not already present; return its index."""
    if class_name not in col_map:
        new_col = NUM_FIXED + len(col_map) + 1
        col_map[class_name] = new_col
        cell = ws.cell(row=1, column=new_col, value=class_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_CLASS
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(new_col)].width = 18
    return col_map[class_name]


# ── Download helpers ───────────────────────────────────────────────────────────

def detect_media_type(url: str) -> str:
    """Return 'image' or 'video' from URL extension or HTTP content-type."""
    suffix = Path(url.split("?")[0]).suffix.lower()
    if suffix in IMAGE_EXTS:
        return "image"
    if suffix in VIDEO_EXTS:
        return "video"
    try:
        r = requests.head(url, timeout=10,
                          headers={"User-Agent": "Mozilla/5.0"},
                          allow_redirects=True)
        ct = r.headers.get("content-type", "")
        if ct.startswith("image/"):
            return "image"
        if ct.startswith("video/"):
            return "video"
    except Exception:
        pass
    return "image"   # safe default


def download_url(url: str, suffix: str) -> Path:
    print("  Downloading…")
    headers = {"User-Agent": "Mozilla/5.0"}
    with requests.get(url, headers=headers, stream=True, timeout=60) as r:
        r.raise_for_status()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        total = 0
        for chunk in r.iter_content(chunk_size=1024 * 1024):
            tmp.write(chunk)
            total += len(chunk)
        tmp.close()
    print(f"  Downloaded {total / 1024 / 1024:.1f} MB → {Path(tmp.name).name}")
    return Path(tmp.name)


# ── Inference helpers ──────────────────────────────────────────────────────────

def run_image_inference(model, path: Path) -> dict:
    """Return {class_name: [conf, conf, …]} for all detections above threshold."""
    results = model.predict(str(path), conf=CONF, iou=IOU, verbose=False)
    detections: dict[str, list] = defaultdict(list)
    for r in results:
        for box in r.boxes:
            cls_name = model.names[int(box.cls)]
            detections[cls_name].append(round(float(box.conf), 4))
    return dict(detections)


def run_video_inference(model, path: Path) -> dict:
    """
    Sample video at VIDEO_FPS frames/sec.
    Returns {class_name: avg_conf_across_frames} (averaged over frames where
    that class appears).
    """
    cap = cv2.VideoCapture(str(path))
    native_fps = cap.get(cv2.CAP_PROP_FPS) or 25
    interval = max(1, int(native_fps / VIDEO_FPS))

    frame_idx = 0
    sampled = 0
    all_confs: dict[str, list] = defaultdict(list)

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        if frame_idx % interval == 0:
            sampled += 1
            results = model.predict(frame, conf=CONF, iou=IOU, verbose=False)
            for r in results:
                for box in r.boxes:
                    cls_name = model.names[int(box.cls)]
                    all_confs[cls_name].append(float(box.conf))
        frame_idx += 1

    cap.release()
    print(f"  Sampled {sampled} frames from {frame_idx} total")

    # Average confidence per class across all frames where it appeared
    return {
        cls: round(sum(confs) / len(confs), 4)
        for cls, confs in all_confs.items()
    }


def save_annotated(model, path: Path, url: str, media_type: str):
    """Save annotated first image / first video frame to OUTPUT_DIR."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stem = Path(url.split("?")[0]).stem[:40] or "result"
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    out  = OUTPUT_DIR / f"{stem}_{ts}_annotated.jpg"

    if media_type == "image":
        results  = model.predict(str(path), conf=CONF, iou=IOU, verbose=False)
        annotated = results[0].plot()
        cv2.imwrite(str(out), annotated)
    else:
        cap = cv2.VideoCapture(str(path))
        ret, frame = cap.read()
        cap.release()
        if ret:
            results  = model.predict(frame, conf=CONF, iou=IOU, verbose=False)
            annotated = results[0].plot()
            cv2.imwrite(str(out), annotated)

    print(f"  Annotated image → {out.name}")


# ── Result classification ──────────────────────────────────────────────────────

def _format_others(detections: dict, exclude: str = "") -> str:
    items = {k: v for k, v in detections.items() if k != exclude}
    if not items:
        return ""
    parts = []
    for k, v in list(items.items())[:8]:
        c = max(v) if isinstance(v, list) else v
        parts.append(f"{k} ({c:.2f})")
    return "Also detected: " + ", ".join(parts)


def compute_result_image(target: str, detections: dict) -> tuple:
    """
    Classify one image result.
    Returns (detected_str, conf_str, result_type, notes).
    detections = {class_name: [conf, …]}
    """
    t_lower = target.strip().lower()
    is_none = t_lower in {"[none]", "none", ""}

    if is_none:
        if detections:
            fp = ", ".join(
                f"{k} ({max(v):.2f})" for k, v in list(detections.items())[:8]
            )
            return "Yes", "N/A", "FP (False Positive)", f"Model detected: {fp}"
        return "No", "N/A", "TN (True Negative)", "Nothing detected as expected."

    matched = next((k for k in detections if k.lower() == t_lower), None)

    if matched:
        confs    = sorted(detections[matched], reverse=True)
        max_conf = confs[0]
        conf_str = ", ".join(f"{c:.2f}" for c in confs)
        rtype    = "TP" if max_conf >= LOW_CONF_THRESH else "TP (Low Conf)"
        notes    = _format_others(detections, exclude=matched)
        return "Yes", conf_str, rtype, notes

    # Not found → FN
    if detections:
        alt = ", ".join(
            f"{k} ({max(v):.2f})" for k, v in list(detections.items())[:5]
        )
        notes = f"Model saw instead: {alt}"
    else:
        notes = "Nothing detected at all."
    return "No", "N/A", "FN (False Negative)", notes


def compute_result_video(target: str, detections: dict) -> tuple:
    """
    Classify one video result.
    detections = {class_name: avg_conf (float)}
    """
    t_lower = target.strip().lower()
    is_none = t_lower in {"[none]", "none", ""}

    if is_none:
        if detections:
            fp = ", ".join(
                f"{k} ({v:.2f})" for k, v in list(detections.items())[:8]
            )
            return "Yes", "N/A", "FP (False Positive)", f"Model detected: {fp}"
        return "No", "N/A", "TN (True Negative)", "Nothing detected as expected."

    matched = next((k for k in detections if k.lower() == t_lower), None)

    if matched:
        avg_conf = detections[matched]
        rtype    = "TP" if avg_conf >= LOW_CONF_THRESH else "TP (Low Conf)"
        notes    = _format_others(detections, exclude=matched)
        return "Yes", f"{avg_conf:.2f} (avg)", rtype, notes

    if detections:
        alt = ", ".join(
            f"{k} ({v:.2f})" for k, v in list(detections.items())[:5]
        )
        notes = f"Model saw instead: {alt}"
    else:
        notes = "Nothing detected at all."
    return "No", "N/A", "FN (False Negative)", notes


# ── Row writing ────────────────────────────────────────────────────────────────

def write_row(ws, col_map: dict, row: int, url: str, target: str,
              detected_str: str, conf_str: str, result_type: str,
              notes: str, detections: dict, media_type: str):
    """Write one result row and colour-code it."""
    ws.cell(row=row, column=1, value=url)
    ws.cell(row=row, column=2, value=target)
    ws.cell(row=row, column=3, value=detected_str)
    ws.cell(row=row, column=4, value=conf_str)
    ws.cell(row=row, column=5, value=result_type)
    ws.cell(row=row, column=6, value=notes if notes else "")

    # Dynamic class columns
    for cls_name, confs in detections.items():
        col_idx = ensure_class_column(ws, col_map, cls_name)
        if media_type == "image":
            val = ", ".join(f"{c:.2f}" for c in
                            sorted(confs, reverse=True))   # list of floats
        else:
            val = f"{confs:.2f} (avg)"   # single float already averaged
        ws.cell(row=row, column=col_idx, value=val)

    # Colour-code the fixed columns
    color = RESULT_COLORS.get(result_type, "FFFFFF")
    fill  = PatternFill("solid", fgColor=color)
    align = Alignment(wrap_text=True, vertical="top")
    for col in range(1, NUM_FIXED + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill  = fill
        cell.alignment = align


# ── Main loop ──────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*62}")
    print("  Hospital V3 Inference — Interactive Loop")
    print(f"  Model : {MODEL_PATH.relative_to(BASE_DIR)}")
    print(f"  Excel : {EXCEL_PATH.relative_to(BASE_DIR)}")
    print(f"  Conf  : {CONF}   IoU: {IOU}   Video FPS: {VIDEO_FPS}")
    print(f"{'='*62}")
    print("  Type 'quit' to exit.\n")

    if not MODEL_PATH.exists():
        print(f"ERROR: model not found at\n  {MODEL_PATH}")
        sys.exit(1)

    print("Loading model…")
    model = YOLO(str(MODEL_PATH))
    print("Model loaded.\n")

    wb, ws, col_map = get_or_create_workbook()

    while True:
        # ── Prompt ────────────────────────────────────────────────────────────
        try:
            url = input("URL (image or video): ").strip()
        except (KeyboardInterrupt, EOFError):
            break
        if not url or url.lower() in STOP_WORDS:
            break

        try:
            target = input("Expected class (or [None] if testing FP): ").strip()
        except (KeyboardInterrupt, EOFError):
            break
        if not target:
            target = "[None]"

        # ── Detect type ────────────────────────────────────────────────────────
        media_type = detect_media_type(url)
        suffix     = ".jpg" if media_type == "image" else ".mp4"
        print(f"  Media type: {media_type}")

        # ── Download ───────────────────────────────────────────────────────────
        try:
            tmp_path = download_url(url, suffix)
        except Exception as e:
            print(f"  Download failed: {e}\n")
            continue

        # ── Inference ──────────────────────────────────────────────────────────
        try:
            if media_type == "image":
                detections = run_image_inference(model, tmp_path)
                detected_str, conf_str, result_type, notes = \
                    compute_result_image(target, detections)
            else:
                detections = run_video_inference(model, tmp_path)
                detected_str, conf_str, result_type, notes = \
                    compute_result_video(target, detections)

            save_annotated(model, tmp_path, url, media_type)

        except Exception as e:
            print(f"  Inference error: {e}\n")
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
            continue
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

        # ── Write Excel row ────────────────────────────────────────────────────
        next_row = ws.max_row + 1
        write_row(ws, col_map, next_row,
                  url, target, detected_str, conf_str,
                  result_type, notes, detections, media_type)
        wb.save(EXCEL_PATH)

        # ── Terminal summary ───────────────────────────────────────────────────
        print(f"\n  ┌─ Result ──────────────────────────────────")
        print(f"  │  Type       : {result_type}")
        print(f"  │  Detected?  : {detected_str}")
        print(f"  │  Confidence : {conf_str}")
        if notes:
            print(f"  │  Notes      : {notes}")
        print(f"  │  Classes detected this run: {len(detections)}")
        if detections:
            # Sort by confidence descending
            def _max_conf(item):
                v = item[1]
                return max(v) if isinstance(v, list) else v

            top = sorted(detections.items(), key=_max_conf, reverse=True)[:8]
            for cls, confs in top:
                c = max(confs) if isinstance(confs, list) else confs
                print(f"  │    {cls:<28} {c:.2f}")
        print(f"  └───────────────────────────────────────────")
        print(f"  Saved → row {next_row} in {EXCEL_PATH.name}\n")

    wb.save(EXCEL_PATH)
    print("Exiting. Excel saved.")


if __name__ == "__main__":
    main()
