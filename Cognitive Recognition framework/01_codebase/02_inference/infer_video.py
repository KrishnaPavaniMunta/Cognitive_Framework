"""
infer_video.py
──────────────
Run inference on a video URL using the trained 89-class model.
Results for target objects are appended to inference_results.xlsx.

Usage:
    python infer_video.py --url <video_url> --target Watch
    python infer_video.py --url <video_url> --target Watch Key --conf 0.3
"""

import argparse
import tempfile
import os
from collections import defaultdict
from pathlib import Path

import cv2
import requests
import openpyxl
from openpyxl import Workbook
from ultralytics import YOLO

BASE_DIR      = Path(__file__).parent.resolve()
MODEL_PATH    = BASE_DIR / "training_output" / "phase1_head" / "weights" / "best.pt"
EXCEL_PATH    = BASE_DIR / "inference_results.xlsx"
OUTPUT_DIR    = BASE_DIR / "inference_output"
MODEL_NAME    = "yolo26n-phase1"

HEADERS = ["Object", "Model", "Lighting", "Viewpoint", "Distance", "Motion",
           "Detection Rate%", "Avg_Confidence"]


def get_or_create_workbook():
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inference Results"
        ws.append(HEADERS)
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="305496")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
    return wb, ws


def download_video(url: str) -> str:
    """Download video from URL to a temp file, return the temp file path."""
    print(f"Downloading video from: {url}")
    headers = {"User-Agent": "Mozilla/5.0"}
    with requests.get(url, headers=headers, stream=True, timeout=60) as r:
        r.raise_for_status()
        suffix = ".mp4"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        total = 0
        for chunk in r.iter_content(chunk_size=1024 * 1024):
            tmp.write(chunk)
            total += len(chunk)
        tmp.close()
    print(f"Downloaded {total / 1024 / 1024:.1f} MB → {tmp.name}")
    return tmp.name


def run_inference(url: str, targets: list, conf: float, iou: float) -> None:
    targets_lower = {t.lower(): t for t in targets}

    print(f"Loading model: {MODEL_PATH}")
    model = YOLO(str(MODEL_PATH))
    class_names = model.names  # {idx: name}

    name_to_idx = {v.lower(): k for k, v in class_names.items()}

    for t in targets:
        if t.lower() not in name_to_idx:
            print(f"  WARNING: '{t}' not found in model classes. Skipping.")
    valid_targets = [t for t in targets if t.lower() in name_to_idx]
    if not valid_targets:
        print("No valid targets found. Exiting.")
        return

    target_idxs = {name_to_idx[t.lower()] for t in valid_targets}

    frames_detected = defaultdict(int)
    conf_sum        = defaultdict(float)
    conf_count      = defaultdict(int)
    prev_centers    = {}
    displacements   = defaultdict(list)
    total_frames    = 0

    print(f"Running inference on: {url}")
    print(f"Target objects: {valid_targets}\n")

    tmp_path = download_video(url)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "result.mp4"

    results = model.predict(
        source=tmp_path,
        conf=conf,
        iou=iou,
        classes=list(target_idxs),
        save=False,
        stream=True,
        verbose=False,
    )

    video_writer = None

    for result in results:
        total_frames += 1
        if total_frames % 50 == 0:
            print(f"  Processed {total_frames} frames…")

        annotated = result.plot()

        # Initialise video writer on first frame
        if video_writer is None:
            h, w = annotated.shape[:2]
            video_writer = cv2.VideoWriter(
                str(out_path),
                cv2.VideoWriter_fourcc(*"mp4v"),
                30,
                (w, h),
            )

        video_writer.write(annotated)

        if result.boxes is None or len(result.boxes) == 0:
            continue

        frame_w = result.orig_shape[1]
        frame_h = result.orig_shape[0]

        seen_in_frame = set()
        for box in result.boxes:
            cls_idx    = int(box.cls[0].item())
            confidence = float(box.conf[0].item())

            seen_in_frame.add(cls_idx)
            conf_sum[cls_idx]   += confidence
            conf_count[cls_idx] += 1

            x1, y1, x2, y2 = box.xyxy[0].tolist()
            cx = ((x1 + x2) / 2) / frame_w
            cy = ((y1 + y2) / 2) / frame_h

            if cls_idx in prev_centers:
                px, py = prev_centers[cls_idx]
                disp = ((cx - px) ** 2 + (cy - py) ** 2) ** 0.5
                displacements[cls_idx].append(disp)
            prev_centers[cls_idx] = (cx, cy)

        for cls_idx in seen_in_frame:
            frames_detected[cls_idx] += 1

    print(f"\nTotal frames processed: {total_frames}")
    if video_writer is not None:
        video_writer.release()
        print(f"Annotated video saved to: {out_path}")
    cv2.destroyAllWindows()
    os.unlink(tmp_path)  # clean up temp file
    if total_frames == 0:
        print("No frames were read. Check the URL.")
        return

    MOTION_THRESHOLD = 0.02  # 2% of normalised frame width

    wb, ws = get_or_create_workbook()

    rows_added = 0
    for target in valid_targets:
        idx      = name_to_idx[target.lower()]
        det_rate = round((frames_detected[idx] / total_frames) * 100, 2)
        avg_conf = round(conf_sum[idx] / conf_count[idx] * 100, 2) if conf_count[idx] else 0.0

        disps = displacements.get(idx, [])
        if disps:
            mean_disp  = sum(disps) / len(disps)
            motion_val = "Moving" if mean_disp > MOTION_THRESHOLD else "Static"
        else:
            motion_val = "Static"

        ws.append([
            target,      # Object
            MODEL_NAME,  # Model
            "",          # Lighting    — fill manually
            "",          # Viewpoint   — fill manually
            "",          # Distance    — fill manually
            motion_val,  # Motion
            det_rate,    # Detection Rate%
            avg_conf,    # Avg_Confidence
        ])
        rows_added += 1
        print(f"  {target:20s} | Detection Rate: {det_rate}%  | Avg Conf: {avg_conf}%  | Motion: {motion_val}")

    wb.save(EXCEL_PATH)
    print(f"\nResults appended to: {EXCEL_PATH}  ({rows_added} row(s) added)")


def main():
    parser = argparse.ArgumentParser(description="YOLO inference on video URL → Excel")
    parser.add_argument("--url",    required=True,             help="Video URL to run inference on")
    parser.add_argument("--target", required=True, nargs="+",  help="Target class name(s), e.g. Watch Key")
    parser.add_argument("--conf",   type=float, default=0.25,  help="Confidence threshold (default: 0.25)")
    parser.add_argument("--iou",    type=float, default=0.45,  help="IoU NMS threshold (default: 0.45)")
    args = parser.parse_args()

    run_inference(args.url, args.target, args.conf, args.iou)


if __name__ == "__main__":
    main()
