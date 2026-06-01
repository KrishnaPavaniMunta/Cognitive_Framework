"""
compare_v1_v3.py
Runs 5 target classes through both V1 and V3 models.
Saves annotated images to outputs/compare_v1_v3/
Logs results to outputs/compare_v1_v3_log.xlsx
"""

import os, tempfile, requests
from pathlib import Path
from collections import defaultdict
import cv2
from ultralytics import YOLO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR   = Path(__file__).parent.resolve()
V1_PATH    = BASE_DIR / "outputs/runs/hospital/phase2_neck_head/weights/best.pt"
V3_PATH    = BASE_DIR / "outputs/runs/hospital_v3/phase2_neck_head/weights/best.pt"
OUT_DIR    = BASE_DIR / "outputs/compare_v1_v3"
EXCEL_PATH = BASE_DIR / "outputs/compare_v1_v3_log.xlsx"
CONF       = 0.10   # low threshold to catch uncertain detections

TESTS = [
    # (target_class, image_url)
    ("person",           "https://images.unsplash.com/photo-1551601651-2a8555f1a136?w=640"),
    ("healthcare_worker","https://images.unsplash.com/photo-1612349317150-e413f6a5b16d?w=640"),
    ("test_tube",        "https://images.unsplash.com/photo-1587691592099-24045742c181?w=640"),
    ("glove",            "https://images.unsplash.com/photo-1583947215259-38e31be8751f?w=640"),
    ("patient_monitor",  "https://images.unsplash.com/photo-1516549655169-df83a0774514?w=640"),
]

def download(url):
    suffix = Path(url.split("?")[0]).suffix or ".jpg"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    r.raise_for_status()
    tmp.write(r.content)
    tmp.close()
    return Path(tmp.name)

def run_model(model, img_path):
    results = model(str(img_path), conf=CONF, iou=0.45, verbose=False)
    r = results[0]
    detections = defaultdict(list)
    if r.boxes is not None:
        names = model.names
        for box in r.boxes:
            cls_name = names[int(box.cls)]
            detections[cls_name].append(round(float(box.conf), 3))
    return r, detections

def save_annotated(result, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    img = result.plot()
    cv2.imwrite(str(path), img)

# ─── Build Excel ───────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1D1E37")
MODEL_FILLS = {
    "V1": PatternFill("solid", fgColor="2E75B6"),
    "V3": PatternFill("solid", fgColor="375623"),
}
HEADER_FONT = Font(bold=True, color="FFFFFF")

wb = Workbook()
ws = wb.active
ws.title = "V1 vs V3 Comparison"

headers = ["Model", "Target Class", "Image URL", "Target Detected?",
           "Target Conf", "All Detections", "Notes"]
widths  = [8, 20, 55, 18, 14, 50, 40]
for ci, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 30

# ─── Main loop ────────────────────────────────────────────────────────────────
print("Loading models...")
v1 = YOLO(str(V1_PATH))
v3 = YOLO(str(V3_PATH))

row = 2
for target, url in TESTS:
    print(f"\n[{target}] Downloading image...")
    try:
        img_path = download(url)
    except Exception as e:
        print(f"  Download failed: {e}")
        continue

    for model_name, model in [("V1", v1), ("V3", v3)]:
        print(f"  Running {model_name}...")
        try:
            result, dets = run_model(model, img_path)
        except Exception as e:
            print(f"  Inference failed: {e}")
            continue

        # Save annotated image
        out_img = OUT_DIR / f"{target}_{model_name}.jpg"
        save_annotated(result, out_img)
        print(f"  Saved: {out_img.name}")

        # Check target
        target_confs = dets.get(target, [])
        detected = "Yes" if target_confs else "No"
        conf_str = ", ".join(str(c) for c in target_confs) if target_confs else "—"
        all_dets = "; ".join(f"{k}:{v}" for k, v in dets.items()) if dets else "nothing"
        notes = f"Also detected: {all_dets}" if dets else "No detections at all"

        fill = MODEL_FILLS[model_name]
        for ci, val in enumerate([model_name, target, url, detected, conf_str, all_dets, notes], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if ci <= 2:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = fill

        # Green/red highlight on "Target Detected?" column
        if target_confs:
            ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="FFC7CE")

        row += 1

    os.unlink(img_path)

wb.save(EXCEL_PATH)
print(f"\nDone.")
print(f"Excel:  {EXCEL_PATH}")
print(f"Images: {OUT_DIR}/")
