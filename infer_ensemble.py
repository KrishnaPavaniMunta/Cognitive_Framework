"""
infer_ensemble.py
Ensembles V1 (106-class hospital) + V3 (109-class) models.
Strategy:
  - Overlapping working classes (wheelchair, door, fire_extinguisher):
      merge boxes from both models via NMS
  - V3-only new classes (bag, exit_sign, spillage): V3 only
  - All other V1 hospital classes: V1 only
Appends to existing outputs/v3_inference_log.xlsx (same sheet as infer_v3.py).
Saves annotated images to outputs/inference_output/ (same folder as infer_v3.py).
"""

import os, tempfile, requests
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import cv2, torch
from ultralytics import YOLO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR   = Path(__file__).parent.resolve()
V1_PATH    = BASE_DIR / "outputs/runs/hospital/phase2_neck_head/weights/best.pt"
V3_PATH    = BASE_DIR / "outputs/runs/hospital_v3/phase2_neck_head/weights/best.pt"
OUT_DIR    = BASE_DIR / "outputs/inference_output"
EXCEL_PATH = BASE_DIR / "outputs/v3_inference_log.xlsx"
OUT_DIR.mkdir(parents=True, exist_ok=True)

CONF            = 0.25
IOU             = 0.45
LOW_CONF_THRESH = 0.50

# Classes where V3 has confirmed working detection — merge V1+V3 boxes via NMS
V3_WORKING_OVERLAP = {"wheelchair", "door", "fire_extinguisher"}
# Classes only in V3 (not in V1 at all)
V3_ONLY_NEW        = {"bag", "exit_sign", "spillage"}


# ── Excel constants (same as infer_v3.py) ─────────────────────────────────────
FIXED_HEADERS = [
    "Image ID",
    "Ground Truth (Target)",
    "Detected?",
    "Confidence Score",
    "Result Type",
    "Notes (False Positives/Misses)",
]
NUM_FIXED    = len(FIXED_HEADERS)
FIXED_WIDTHS = {"A": 55, "B": 24, "C": 12, "D": 22, "E": 22, "F": 60}

RESULT_COLORS = {
    "TP":                  "C6EFCE",
    "TP (Low Conf)":       "FFEB9C",
    "FN (False Negative)": "FFC7CE",
    "FP (False Positive)": "FFCC99",
    "TN (True Negative)":  "D9E1F2",
}
HEADER_FILL_FIXED = PatternFill("solid", fgColor="1D1E37")
HEADER_FILL_CLASS = PatternFill("solid", fgColor="FA643F")
HEADER_FONT       = Font(bold=True, color="FFFFFF")


# ── NMS helper ────────────────────────────────────────────────────────────────
def nms_merge(combined, iou_thresh=0.45):
    """combined: list of (x1,y1,x2,y2,conf). Returns filtered list."""
    if not combined:
        return []
    boxes  = torch.tensor([[d[0], d[1], d[2], d[3]] for d in combined], dtype=torch.float32)
    scores = torch.tensor([d[4] for d in combined], dtype=torch.float32)
    from torchvision.ops import nms
    kept = nms(boxes, scores, iou_thresh).tolist()
    return [combined[i] for i in kept]


# ── Ensemble inference ────────────────────────────────────────────────────────
def ensemble_infer(v1, v3, img_path):
    """Returns dict: {class_name: [(x1,y1,x2,y2,conf), ...]}"""
    r1 = v1(str(img_path), conf=CONF, iou=IOU, verbose=False)[0]
    r3 = v3(str(img_path), conf=CONF, iou=IOU, verbose=False)[0]

    v1_dets = defaultdict(list)
    if r1.boxes is not None:
        for box in r1.boxes:
            name = v1.names[int(box.cls)]
            xyxy = box.xyxy[0].cpu().tolist()
            conf = float(box.conf)
            v1_dets[name].append((*xyxy, conf))

    v3_dets = defaultdict(list)
    if r3.boxes is not None:
        for box in r3.boxes:
            name = v3.names[int(box.cls)]
            xyxy = box.xyxy[0].cpu().tolist()
            conf = float(box.conf)
            v3_dets[name].append((*xyxy, conf))

    merged = {}

    # 1. V3-only new classes
    for cls in V3_ONLY_NEW:
        if cls in v3_dets:
            merged[cls] = v3_dets[cls]

    # 2. Overlapping classes — merge then NMS
    for cls in V3_WORKING_OVERLAP:
        combined = v1_dets.get(cls, []) + v3_dets.get(cls, [])
        if combined:
            merged[cls] = nms_merge(combined, IOU)

    # 3. V1-only classes
    exclude = V3_WORKING_OVERLAP | V3_ONLY_NEW
    for cls, dets in v1_dets.items():
        if cls not in exclude:
            merged[cls] = dets

    return merged


# ── Draw merged detections (high-visibility boxes) ────────────────────────────
def draw_merged(img_path, merged_dets):
    img = cv2.imread(str(img_path))
    h_img, w_img = img.shape[:2]
    import random; rng = random.Random(42)
    colors = {}
    for cls, dets in merged_dets.items():
        if cls not in colors:
            colors[cls] = (rng.randint(50, 255), rng.randint(50, 255), rng.randint(50, 255))
        c = colors[cls]
        for (x1, y1, x2, y2, conf) in dets:
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
            # Thick border box + inner white outline for contrast
            cv2.rectangle(img, (x1, y1), (x2, y2), (255, 255, 255), 5)
            cv2.rectangle(img, (x1, y1), (x2, y2), c, 3)
            # Label background + text
            label = f"{cls} {conf:.2f}"
            scale = max(0.6, min(1.2, w_img / 800))
            thickness = max(2, int(scale * 2))
            (tw, th), baseline = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, scale, thickness)
            ty = max(y1 - 4, th + 4)
            cv2.rectangle(img, (x1, ty - th - 4), (x1 + tw + 4, ty + baseline), (0, 0, 0), -1)
            cv2.putText(img, label, (x1 + 2, ty - 2),
                        cv2.FONT_HERSHEY_SIMPLEX, scale, (255, 255, 255), thickness)
    return img


# ── Workbook helpers (same as infer_v3.py) ────────────────────────────────────
def get_or_create_workbook():
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        col_map = {}
        for cell in ws[1]:
            if cell.column > NUM_FIXED and cell.value:
                col_map[str(cell.value)] = cell.column
        return wb, ws, col_map

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "V3 Inference Log"
    for ci, h in enumerate(FIXED_HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_FIXED
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    for col, w in FIXED_WIDTHS.items():
        ws.column_dimensions[col].width = w
    return wb, ws, {}


def ensure_class_column(ws, col_map, class_name):
    if class_name not in col_map:
        new_col = NUM_FIXED + len(col_map) + 1
        col_map[class_name] = new_col
        cell = ws.cell(row=1, column=new_col, value=class_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_CLASS
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(new_col)].width = 18
    return col_map[class_name]


def write_row(ws, col_map, row_idx, url, target, detected_str,
              conf_str, result_type, notes, flat_dets):
    ws.cell(row=row_idx, column=1, value=url)
    ws.cell(row=row_idx, column=2, value=target)
    ws.cell(row=row_idx, column=3, value=detected_str)
    ws.cell(row=row_idx, column=4, value=conf_str)
    ws.cell(row=row_idx, column=5, value=result_type)
    ws.cell(row=row_idx, column=6, value=notes or "")
    for cls_name, confs in flat_dets.items():
        col_idx = ensure_class_column(ws, col_map, cls_name)
        ws.cell(row=row_idx, column=col_idx,
                value=", ".join(f"{c:.2f}" for c in sorted(confs, reverse=True)))
    color = RESULT_COLORS.get(result_type, "FFFFFF")
    fill  = PatternFill("solid", fgColor=color)
    align = Alignment(wrap_text=True, vertical="top")
    for col in range(1, NUM_FIXED + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.alignment = align


# ── Result classification ─────────────────────────────────────────────────────
def classify_result(target, flat_dets):
    """flat_dets = {cls: [conf, ...]}. Returns (detected_str, conf_str, result_type, notes)."""
    t_lower = target.strip().lower()
    is_none = t_lower in {"[none]", "none", ""}
    if is_none:
        if flat_dets:
            fp = ", ".join(f"{k} ({max(v):.2f})" for k, v in list(flat_dets.items())[:8])
            return "Yes", "N/A", "FP (False Positive)", f"Model detected: {fp}"
        return "No", "N/A", "TN (True Negative)", "Nothing detected as expected."
    matched = next((k for k in flat_dets if k.lower() == t_lower), None)
    if matched:
        confs    = sorted(flat_dets[matched], reverse=True)
        conf_str = ", ".join(f"{c:.2f}" for c in confs)
        rtype    = "TP" if confs[0] >= LOW_CONF_THRESH else "TP (Low Conf)"
        others   = {k: v for k, v in flat_dets.items() if k != matched}
        notes    = ("Also detected: " + ", ".join(
            f"{k} ({max(v):.2f})" for k, v in list(others.items())[:8])) if others else ""
        return "Yes", conf_str, rtype, notes
    if flat_dets:
        alt = ", ".join(f"{k} ({max(v):.2f})" for k, v in list(flat_dets.items())[:5])
        notes = f"Model saw instead: {alt}"
    else:
        notes = "Nothing detected at all."
    return "No", "N/A", "FN (False Negative)", notes


# ── Download helper ───────────────────────────────────────────────────────────
def download(url):
    suffix = Path(url.split("?")[0]).suffix or ".jpg"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    r.raise_for_status()
    tmp.write(r.content)
    tmp.close()
    return Path(tmp.name)


def source_label(cls):
    if cls in V3_ONLY_NEW:        return "V3"
    if cls in V3_WORKING_OVERLAP: return "V1+V3"
    return "V1"


# ── Main ──────────────────────────────────────────────────────────────────────
print("Loading V1 (106-class hospital)...")
v1 = YOLO(str(V1_PATH))
print("Loading V3 (109-class)...")
v3 = YOLO(str(V3_PATH))
print("\nEnsemble ready.")
print(f"  Overlap (V1+V3 NMS) : {sorted(V3_WORKING_OVERLAP)}")
print(f"  V3-only new classes : {sorted(V3_ONLY_NEW)}")
print(f"  All other classes   : V1 only")
print(f"  Logging to          : {EXCEL_PATH.relative_to(BASE_DIR)}")
print("\nPaste an image URL (or 'quit' to exit).\n")

wb, ws, col_map = get_or_create_workbook()
excel_row = ws.max_row + 1
img_counter = 1

while True:
    url = input("URL: ").strip()
    if url.lower() in {"quit", "exit", "q", "stop"}:
        break
    if not url:
        continue

    target = input("Expected class (or [None]): ").strip() or "[None]"

    try:
        img_path = download(url)
        print("  Downloaded.")
    except Exception as e:
        print(f"  Download failed: {e}")
        continue

    try:
        merged = ensemble_infer(v1, v3, img_path)
    except Exception as e:
        print(f"  Inference failed: {e}")
        os.unlink(img_path)
        continue

    # Build flat detections dict for Excel + classify
    flat_dets = {cls: [d[4] for d in dets] for cls, dets in merged.items()}

    # Print results
    if merged:
        print("  Detections:")
        for cls in sorted(merged):
            confs = [round(d[4], 3) for d in merged[cls]]
            print(f"    [{source_label(cls)}] {cls}: {confs}")
    else:
        print("  No detections.")

    detected_str, conf_str, result_type, notes = classify_result(target, flat_dets)
    # Prepend source model info to notes when target was found
    if detected_str == "Yes" and target.lower() not in {"[none]", "none", ""}:
        tgt_key = next((k for k in flat_dets if k.lower() == target.strip().lower()), None)
        src = f"[{source_label(tgt_key)}] " if tgt_key else ""
        notes = src + notes if notes else src.rstrip()
    print(f"  Result: {result_type}  |  conf: {conf_str}")

    # Save annotated image
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = Path(url.split("?")[0]).stem[:35] or "img"
    out_name = OUT_DIR / f"{stem}_{ts}_ensemble.jpg"
    ann = draw_merged(img_path, merged)
    cv2.imwrite(str(out_name), ann)
    print(f"  Saved: {out_name.name}")

    write_row(ws, col_map, excel_row, url, target, detected_str,
              conf_str, result_type, notes, flat_dets)
    wb.save(EXCEL_PATH)

    excel_row += 1
    img_counter += 1
    os.unlink(img_path)

wb.save(EXCEL_PATH)
print(f"\nDone. Excel: {EXCEL_PATH}")
